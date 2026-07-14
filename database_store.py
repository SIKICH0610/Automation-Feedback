from __future__ import annotations

from contextlib import contextmanager
from copy import copy
from datetime import date, datetime, time
import hashlib
import json
from pathlib import Path
import re
import sqlite3
import threading
from typing import Any, Iterator
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


SCHEMA_VERSION = "1"
BOOLEAN_COLUMNS = {
    "Group Chat",
    "Before Class Informing",
    "Class Attendance",
    "Pre-quiz informing",
}
LONG_TEXT_COLUMNS = {
    "Remark for Student",
    "Additional Comment",
    "Homework Reflection",
    "Quiz Feedback",
    "Second Feeback",
    "Feedback",
    "Send Error",
    "Quiz1 Feedback",
    "Quiz1 Mistake",
    "Quiz2 Feedback",
    "Quiz2 Mistake",
}
COLUMN_OPTIONS = {
    "Group Chat": ["", "TRUE", "FALSE"],
    "Before Class Informing": ["", "TRUE", "FALSE"],
    "Pre-quiz informing": ["", "TRUE", "FALSE"],
    "Parent Language": ["", "Chinese", "English"],
    "Preferred Channel": ["", "wecom", "whatsapp"],
    "WhatsApp Target Type": ["", "group_search", "phone"],
    "Class Attendance": ["", "TRUE", "FALSE", "Present", "Absent"],
    "Note-taking": ["", "Excellent", "Good", "Needs Reminder", "Not Observed"],
    "Listening & Focus": [
        "",
        "Very Focused",
        "Focused",
        "Sometimes Distracted",
        "Sometimes Needs Reminder",
        "Needs Support",
        "Not Observed",
    ],
    "Participation": [
        "",
        "Very Active",
        "Active",
        "Helpful",
        "Asks for Help",
        "Average",
        "Needs Encouragement",
        "Not Observed",
    ],
    "Practice Speed & Accuracy": [
        "",
        "Excellent",
        "High Accuracy",
        "Fast but Needs Care",
        "Good but Rushed",
        "Good",
        "Normal",
        "Slow but Thoughtful",
        "Needs Support",
        "Not Observed",
    ],
    "Proof Logic": ["", "Strong", "Good with Hint", "Needs Practice", "Not Observed"],
    "Calculation": ["", "Good", "Needs More Care", "Not Observed"],
}
STUDENT_COLUMNS = {
    "First Name",
    "Last Name",
    "uid",
    "Group Chat",
    "Parent Language",
    "Preferred Channel",
    "WhatsApp Phone",
    "WhatsApp Search Key",
    "WhatsApp Target Type",
}
AUDIT_COLUMNS = {"Send Status", "Send Error", "Last Attempt"}


class StoreError(Exception):
    pass


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def announcement_filename(sheet_name: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", sheet_name).strip("._")
    if not stem:
        digest = hashlib.sha1(sheet_name.encode("utf-8")).hexdigest()[:8]
        stem = f"class_{digest}"
    return f"{stem}.txt"


def column_group(header: str) -> str:
    lowered = header.lower().replace(" ", "")
    if header in STUDENT_COLUMNS:
        return "student"
    if header in AUDIT_COLUMNS:
        return "audit"
    if lowered.startswith("quiz1"):
        return "quiz1"
    if lowered.startswith("quiz2"):
        return "quiz2"
    return "general"


def coerce_cell_value(header: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if header == "uid":
        return str(value).strip()
    if header in BOOLEAN_COLUMNS and isinstance(value, str):
        normalized = value.strip().lower()
        if normalized == "true":
            return True
        if normalized == "false":
            return False
    return value


def header_cells(worksheet: Any) -> list[tuple[int, str]]:
    headers: list[tuple[int, str]] = []
    seen: set[str] = set()
    for column in range(1, worksheet.max_column + 1):
        raw_header = worksheet.cell(1, column).value
        if raw_header is None or not str(raw_header).strip():
            continue
        header = str(raw_header).strip()
        if header in seen:
            raise StoreError(f"Duplicate column header {header!r} in {worksheet.title!r}.")
        seen.add(header)
        headers.append((column, header))
    return headers


class SQLiteFeedbackStore:
    def __init__(
        self,
        *,
        database_path: Path,
        source_workbook_path: Path,
        announcement_dir: Path,
        app_data_dir: Path,
        export_dir: Path,
    ) -> None:
        self.database_path = database_path.resolve()
        self.source_workbook_path = source_workbook_path.resolve()
        self.announcement_dir = announcement_dir.resolve()
        self.app_data_dir = app_data_dir.resolve()
        self.export_dir = export_dir.resolve()
        self.template_path = self.app_data_dir / "workbook_template.xlsx"
        self.workbook_path = self.app_data_dir / "runtime_workbook.xlsx"
        self.lock = threading.RLock()

        self.app_data_dir.mkdir(parents=True, exist_ok=True)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._initialize_database()
        self.ensure_announcement_files()

    @contextmanager
    def _connect(self) -> Iterator[sqlite3.Connection]:
        connection = sqlite3.connect(self.database_path, timeout=30)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        try:
            yield connection
        except Exception:
            connection.rollback()
            raise
        else:
            connection.commit()
        finally:
            connection.close()

    def _initialize_database(self) -> None:
        new_database = not self.database_path.exists()
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS classes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    position INTEGER NOT NULL
                );

                CREATE TABLE IF NOT EXISTS columns (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    position INTEGER NOT NULL,
                    width REAL NOT NULL DEFAULT 13,
                    UNIQUE(class_id, name),
                    FOREIGN KEY(class_id) REFERENCES classes(id) ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS students (
                    id TEXT PRIMARY KEY,
                    class_id INTEGER NOT NULL,
                    position INTEGER NOT NULL,
                    values_json TEXT NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(class_id) REFERENCES classes(id) ON DELETE CASCADE
                );

                CREATE INDEX IF NOT EXISTS students_class_position
                ON students(class_id, position);
                """
            )
            class_count = connection.execute("SELECT COUNT(*) FROM classes").fetchone()[0]
            if class_count == 0:
                if not self.source_workbook_path.exists():
                    raise FileNotFoundError(
                        "No SQLite data exists yet and the import workbook was not found at "
                        f"{self.source_workbook_path}"
                    )
                self._import_workbook(connection)
                self._create_blank_template()
            elif not self.template_path.exists():
                self._create_template_from_database(connection)

            connection.execute(
                "INSERT OR REPLACE INTO meta(key, value) VALUES('schema_version', ?)",
                (SCHEMA_VERSION,),
            )
            if new_database:
                connection.execute(
                    "INSERT OR REPLACE INTO meta(key, value) VALUES('import_source', ?)",
                    (str(self.source_workbook_path),),
                )

    def _import_workbook(self, connection: sqlite3.Connection) -> None:
        workbook = load_workbook(self.source_workbook_path, data_only=False)
        try:
            for sheet_position, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                cursor = connection.execute(
                    "INSERT INTO classes(name, position) VALUES(?, ?)",
                    (sheet_name, sheet_position),
                )
                class_id = int(cursor.lastrowid)
                headers = header_cells(worksheet)
                for position, (column_index, header) in enumerate(headers):
                    width = worksheet.column_dimensions[get_column_letter(column_index)].width or 13
                    connection.execute(
                        "INSERT INTO columns(class_id, name, position, width) VALUES(?, ?, ?, ?)",
                        (class_id, header, position, float(width)),
                    )

                student_position = 0
                for row_number in range(2, worksheet.max_row + 1):
                    values = {
                        header: json_value(worksheet.cell(row_number, column_index).value)
                        for column_index, header in headers
                    }
                    if not any(value not in (None, "") for value in values.values()):
                        continue
                    connection.execute(
                        """
                        INSERT INTO students(id, class_id, position, values_json)
                        VALUES(?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            class_id,
                            student_position,
                            json.dumps(values, ensure_ascii=False),
                        ),
                    )
                    student_position += 1
        finally:
            workbook.close()

    def _create_blank_template(self) -> None:
        workbook = load_workbook(self.source_workbook_path)
        try:
            for worksheet in workbook.worksheets:
                for row_number in range(2, worksheet.max_row + 1):
                    for column_number in range(1, worksheet.max_column + 1):
                        worksheet.cell(row_number, column_number).value = None
            temporary = self.template_path.with_suffix(".tmp.xlsx")
            workbook.save(temporary)
            temporary.replace(self.template_path)
        finally:
            workbook.close()

    def _create_template_from_database(self, connection: sqlite3.Connection) -> None:
        workbook = Workbook()
        try:
            class_rows = connection.execute(
                "SELECT id, name FROM classes ORDER BY position"
            ).fetchall()
            for index, class_row in enumerate(class_rows):
                worksheet = workbook.active if index == 0 else workbook.create_sheet()
                worksheet.title = str(class_row["name"])
                columns = self._columns(connection, int(class_row["id"]))
                for column_index, column in enumerate(columns, start=1):
                    worksheet.cell(1, column_index).value = str(column["name"])
                    worksheet.column_dimensions[get_column_letter(column_index)].width = float(
                        column["width"]
                    )
                worksheet.freeze_panes = "A2"
                if columns:
                    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

            temporary = self.template_path.with_suffix(".tmp.xlsx")
            workbook.save(temporary)
            temporary.replace(self.template_path)
        finally:
            workbook.close()

    def _class_row(self, connection: sqlite3.Connection, sheet_name: str) -> sqlite3.Row:
        row = connection.execute(
            "SELECT id, name, position FROM classes WHERE name = ?",
            (sheet_name,),
        ).fetchone()
        if row is None:
            raise StoreError(f"Sheet {sheet_name!r} was not found.")
        return row

    def sheet_names(self) -> list[str]:
        with self.lock, self._connect() as connection:
            rows = connection.execute(
                "SELECT name FROM classes ORDER BY position"
            ).fetchall()
            return [str(row["name"]) for row in rows]

    def announcement_path(self, sheet_name: str) -> Path:
        if sheet_name not in self.sheet_names():
            raise StoreError(f"Sheet {sheet_name!r} was not found.")
        return self.announcement_dir / announcement_filename(sheet_name)

    def ensure_announcement_files(self) -> None:
        self.announcement_dir.mkdir(parents=True, exist_ok=True)
        for sheet_name in self.sheet_names():
            path = self.announcement_dir / announcement_filename(sheet_name)
            path.touch(exist_ok=True)

    def read_announcement(self, sheet_name: str) -> tuple[str, Path]:
        path = self.announcement_path(sheet_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.touch(exist_ok=True)
        return path.read_text(encoding="utf-8"), path

    def write_announcement(self, sheet_name: str, text: str) -> Path:
        path = self.announcement_path(sheet_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text.strip(), encoding="utf-8")
        return path

    @staticmethod
    def _columns(connection: sqlite3.Connection, class_id: int) -> list[sqlite3.Row]:
        return connection.execute(
            """
            SELECT name, position, width
            FROM columns
            WHERE class_id = ?
            ORDER BY position
            """,
            (class_id,),
        ).fetchall()

    @staticmethod
    def _students(connection: sqlite3.Connection, class_id: int) -> list[sqlite3.Row]:
        return connection.execute(
            """
            SELECT id, position, values_json
            FROM students
            WHERE class_id = ?
            ORDER BY position
            """,
            (class_id,),
        ).fetchall()

    def load_sheet(self, sheet_name: str) -> dict[str, Any]:
        with self.lock, self._connect() as connection:
            class_row = self._class_row(connection, sheet_name)
            class_id = int(class_row["id"])
            columns: list[dict[str, Any]] = []
            for column in self._columns(connection, class_id):
                header = str(column["name"])
                options = COLUMN_OPTIONS.get(header)
                columns.append(
                    {
                        "key": header,
                        "label": header,
                        "group": column_group(header),
                        "kind": "select" if options else (
                            "long_text" if header in LONG_TEXT_COLUMNS else "text"
                        ),
                        "options": options or [],
                        "width": min(max(round(float(column["width"]) * 7.2), 110), 360),
                    }
                )

            rows: list[dict[str, Any]] = []
            for position, student in enumerate(self._students(connection, class_id)):
                values = json.loads(str(student["values_json"]))
                rows.append(
                    {
                        "student_id": str(student["id"]),
                        "excel_row": position + 2,
                        "values": values,
                    }
                )

        announcement, announcement_path = self.read_announcement(sheet_name)
        return {
            "sheet": sheet_name,
            "columns": columns,
            "rows": rows,
            "announcement": announcement,
            "announcement_path": str(announcement_path),
        }

    def search_students(self, query: str, limit: int = 50) -> list[dict[str, Any]]:
        normalized_query = query.strip().casefold()
        if not normalized_query:
            return []

        matches: list[tuple[tuple[int, int, int], dict[str, Any]]] = []
        with self.lock, self._connect() as connection:
            classes = connection.execute(
                "SELECT id, name, position FROM classes ORDER BY position"
            ).fetchall()
            for class_row in classes:
                class_id = int(class_row["id"])
                for student in self._students(connection, class_id):
                    values = json.loads(str(student["values_json"]))
                    first_name = str(values.get("First Name") or "").strip()
                    last_name = str(values.get("Last Name") or "").strip()
                    uid = str(values.get("uid") or "").strip()
                    full_name = " ".join(part for part in (first_name, last_name) if part)
                    searchable = (first_name, last_name, full_name, uid)
                    if not any(normalized_query in value.casefold() for value in searchable):
                        continue

                    exact_uid = 0 if uid.casefold() == normalized_query else 1
                    name_prefix = 0 if full_name.casefold().startswith(normalized_query) else 1
                    result = {
                        "student_id": str(student["id"]),
                        "sheet": str(class_row["name"]),
                        "excel_row": int(student["position"]) + 2,
                        "first_name": first_name,
                        "last_name": last_name,
                        "full_name": full_name,
                        "uid": uid,
                    }
                    matches.append(
                        (
                            (
                                exact_uid,
                                name_prefix,
                                int(class_row["position"]) * 10000
                                + int(student["position"]),
                            ),
                            result,
                        )
                    )

        matches.sort(key=lambda item: item[0])
        safe_limit = max(1, min(int(limit), 100))
        return [result for _, result in matches[:safe_limit]]

    def save_sheet(self, sheet_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
        if not isinstance(rows, list):
            raise StoreError("Rows must be provided as a list.")

        with self.lock, self._connect() as connection:
            class_row = self._class_row(connection, sheet_name)
            class_id = int(class_row["id"])
            known_headers = {
                str(row["name"]) for row in self._columns(connection, class_id)
            }
            existing_ids = {
                str(row["id"]) for row in self._students(connection, class_id)
            }
            submitted_ids: set[str] = set()
            for item in rows:
                if not isinstance(item, dict) or not isinstance(item.get("values"), dict):
                    raise StoreError("Each row must contain a values object.")
                student_id = str(item.get("student_id") or "").strip()
                if not student_id:
                    continue
                if student_id in submitted_ids:
                    raise StoreError(f"Student record {student_id!r} was submitted twice.")
                if student_id not in existing_ids:
                    raise StoreError(f"Student record {student_id!r} was not found.")
                submitted_ids.add(student_id)

            missing_ids = existing_ids - submitted_ids
            if missing_ids:
                raise StoreError(
                    "The roster save was incomplete. Reload the class sheet and try again."
                )

            for position, item in enumerate(rows):
                if not isinstance(item, dict) or not isinstance(item.get("values"), dict):
                    raise StoreError("Each row must contain a values object.")
                values = item["values"]
                unknown_headers = set(values) - known_headers
                if unknown_headers:
                    raise StoreError(f"Unknown columns: {', '.join(sorted(unknown_headers))}")

                normalized_values = {
                    header: coerce_cell_value(header, values.get(header))
                    for header in known_headers
                }
                student_id = str(item.get("student_id") or "").strip()
                if student_id:
                    if student_id not in existing_ids:
                        raise StoreError(f"Student record {student_id!r} was not found.")
                    connection.execute(
                        """
                        UPDATE students
                        SET position = ?, values_json = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ? AND class_id = ?
                        """,
                        (
                            position,
                            json.dumps(normalized_values, ensure_ascii=False),
                            student_id,
                            class_id,
                        ),
                    )
                else:
                    student_id = str(uuid4())
                    connection.execute(
                        """
                        INSERT INTO students(id, class_id, position, values_json)
                        VALUES(?, ?, ?, ?)
                        """,
                        (
                            student_id,
                            class_id,
                            position,
                            json.dumps(normalized_values, ensure_ascii=False),
                        ),
                    )
                    existing_ids.add(student_id)

        return self.load_sheet(sheet_name)

    @staticmethod
    def _copy_row_format(worksheet: Any, source_row: int, target_row: int) -> None:
        if source_row < 2:
            return
        if worksheet.row_dimensions[source_row].height is not None:
            worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(source_row, column)
            target = worksheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format

    def _write_database_to_workbook(self, output_path: Path) -> dict[str, dict[str, int]]:
        workbook = load_workbook(self.template_path)
        row_maps: dict[str, dict[str, int]] = {}
        try:
            with self._connect() as connection:
                for sheet_name in self.sheet_names():
                    class_row = self._class_row(connection, sheet_name)
                    class_id = int(class_row["id"])
                    worksheet = workbook[sheet_name]
                    columns = self._columns(connection, class_id)
                    students = self._students(connection, class_id)

                    for column_index, column in enumerate(columns, start=1):
                        worksheet.cell(1, column_index).value = str(column["name"])

                    required_last_row = len(students) + 1
                    if required_last_row > worksheet.max_row:
                        source_row = 2 if worksheet.max_row >= 2 else 1
                        for target_row in range(worksheet.max_row + 1, required_last_row + 1):
                            self._copy_row_format(worksheet, source_row, target_row)

                    row_map: dict[str, int] = {}
                    for position, student in enumerate(students):
                        row_number = position + 2
                        row_map[str(student["id"])] = row_number
                        values = json.loads(str(student["values_json"]))
                        for column_index, column in enumerate(columns, start=1):
                            header = str(column["name"])
                            worksheet.cell(row_number, column_index).value = values.get(header)

                    clear_from = len(students) + 2
                    for row_number in range(clear_from, worksheet.max_row + 1):
                        for column_index in range(1, worksheet.max_column + 1):
                            worksheet.cell(row_number, column_index).value = None
                    row_maps[sheet_name] = row_map

            output_path.parent.mkdir(parents=True, exist_ok=True)
            temporary = output_path.with_suffix(".tmp.xlsx")
            workbook.save(temporary)
            temporary.replace(output_path)
            return row_maps
        finally:
            workbook.close()

    def prepare_runtime_workbook(self) -> dict[str, dict[str, int]]:
        with self.lock:
            return self._write_database_to_workbook(self.workbook_path)

    def sync_runtime_workbook(self) -> None:
        if not self.workbook_path.exists():
            return

        with self.lock:
            workbook = load_workbook(self.workbook_path, data_only=False)
            try:
                with self._connect() as connection:
                    for sheet_name in self.sheet_names():
                        if sheet_name not in workbook.sheetnames:
                            continue
                        class_row = self._class_row(connection, sheet_name)
                        class_id = int(class_row["id"])
                        worksheet = workbook[sheet_name]
                        headers = header_cells(worksheet)
                        existing_columns = {
                            str(row["name"]): row for row in self._columns(connection, class_id)
                        }
                        for position, (column_index, header) in enumerate(headers):
                            if header not in existing_columns:
                                width = (
                                    worksheet.column_dimensions[get_column_letter(column_index)].width
                                    or 13
                                )
                                connection.execute(
                                    """
                                    INSERT INTO columns(class_id, name, position, width)
                                    VALUES(?, ?, ?, ?)
                                    """,
                                    (class_id, header, position, float(width)),
                                )

                        students = self._students(connection, class_id)
                        for position, student in enumerate(students):
                            row_number = position + 2
                            values = {
                                header: json_value(worksheet.cell(row_number, column_index).value)
                                for column_index, header in headers
                            }
                            connection.execute(
                                """
                                UPDATE students
                                SET values_json = ?, updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                                """,
                                (
                                    json.dumps(values, ensure_ascii=False),
                                    str(student["id"]),
                                ),
                            )
            finally:
                workbook.close()

    def remove_runtime_workbook(self) -> None:
        try:
            self.workbook_path.unlink(missing_ok=True)
        except OSError:
            pass

    def export_public_workbook(self) -> Path:
        with self.lock:
            output_path = self.export_dir / "Student_Feedback_Export.xlsx"
            self._write_database_to_workbook(output_path)
            return output_path



