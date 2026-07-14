from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from openai_api import DEFAULT_OPENAI_MODEL, create_response


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = PROJECT_DIR / "Geo_TTh_Student_Script_fixed_rows_only.xlsx"
DEFAULT_SHEET = "Geo TTh"
ADDITIONAL_COMMENT_COLUMN = "Additional Comment"
SECOND_QUIZ_COLUMNS = ("Quiz2 Score", "Second Quiz Score")
FIRST_QUIZ_SCORE_COLUMNS = (
    "Quiz1 Score",
    "First Quiz Score",
    "Quiz Feedback",
    "First Quiz Feedback",
    "Quiz Score",
    "Second Feedback",
    "Second Feeback",
)
QUIZ_AVERAGE_COLUMNS = ("Quiz1 Average", "Quiz Average", "Class Quiz Average", "First Quiz Average", "Average Quiz Score")
SECOND_QUIZ_AVERAGE_COLUMNS = ("Quiz2 Average", "Second Quiz Average")
QUIZ_BANK_COLUMNS = ("Quiz Bank", "Quiz Type", "Quiz Name")
QUIZ1_MISTAKE_COLUMNS = ("Quiz1 Mistake", "Quiz1 Mistakes", "First Quiz Mistake", "First Quiz Mistakes")
QUIZ2_MISTAKE_COLUMNS = ("Quiz2 Mistake", "Quiz2 Mistakes", "Second Quiz Mistake", "Second Quiz Mistakes")
QUIZ_MISTAKE_COLUMNS = (
    "Quiz Mistake Questions",
    "Mistake Questions",
    "Mistake Question",
    "Wrong Questions",
    "Incorrect Questions",
    "Quiz Mistakes",
)
FEEDBACK_TYPE_CHOICES = ("comprehensive", "general", "quiz")


PARENT_COMMENT_STYLE_GUIDE = """
Write like a real teacher leaving a thoughtful parent update.
Blend quiz performance, classroom behavior, and next steps into natural paragraphs.
Do not use label-style sections, checklist wording, or repeated sentence frames.
Avoid structures like "Regarding the student's performance" or "The main suggestions are".
When several areas need work, vary the phrasing so it sounds spoken rather than copied from a template.
Use natural sentence breaks instead of colons or semicolons.
End with the standard group-chat question sentence.
""".strip()

ZH_PARENT_COMMENT_STYLE_GUIDE = """
中文家长反馈要像老师真实写给家长的消息。
开头可以保留“家长您好”，之后不要反复说“家长”，统一用“您”。
不要写成“关于某某的课堂表现”或“建议后续关注”这种模板句。
如果有多个需要改进的地方，第二个及之后可以自然加入“也”，例如“计算细节也需要更加仔细”。
语气要具体、温和、顺口，像在群里发给家长的说明。
""".strip()

EN_PARENT_COMMENT_STYLE_GUIDE = """
English parent feedback should sound like a natural teacher update.
Do not use report-style labels or repeated template openings.
Connect the quiz result, class habits, and next step in a smooth paragraph.
Keep the tone warm, direct, and parent-friendly.
""".strip()


EN_PHRASES = {
    "Note-taking": {
        "Excellent": "took excellent notes",
        "Good": "took good notes",
        "Needs Reminder": "needed reminders to keep notes complete",
    },
    "Listening & Focus": {
        "Very Focused": "stayed very focused",
        "Focused": "stayed focused",
        "Sometimes Distracted": "was sometimes distracted",
        "Sometimes Needs Reminder": "needed occasional reminders to stay focused",
        "Needs Support": "needed support staying focused",
    },
    "Participation": {
        "Very Active": "participated very actively",
        "Active": "participated actively",
        "Helpful": "helped classmates during practice",
        "Asks for Help": "asked for help when needed",
        "Average": "participated at an average level",
        "Needs Encouragement": "would benefit from more encouragement to participate",
    },
    "Practice Speed & Accuracy": {
        "Excellent": "worked through practice problems with excellent speed and accuracy",
        "High Accuracy": "showed high accuracy on practice problems",
        "Fast but Needs Care": "worked quickly and should continue checking details carefully",
        "Good but Rushed": "did well but sometimes rushed",
        "Good": "did well on practice problems",
        "Normal": "worked at a steady pace",
        "Slow but Thoughtful": "worked slowly but thoughtfully",
        "Needs Support": "needed support with practice problems",
    },
    "Proof Logic": {
        "Strong": "showed strong proof logic",
        "Good with Hint": "understood proof logic well with hints",
        "Needs Practice": "needs more practice with proof logic",
    },
    "Calculation": {
        "Good": "calculated accurately",
        "Needs More Care": "should take more care with calculations",
    },
}


ZH_PHRASES = {
    "Note-taking": {
        "Excellent": "笔记非常完整",
        "Good": "笔记比较完整",
        "Needs Reminder": "笔记方面需要提醒",
    },
    "Listening & Focus": {
        "Very Focused": "听课非常专注",
        "Focused": "听课比较专注",
        "Sometimes Distracted": "有时会分心",
        "Sometimes Needs Reminder": "偶尔需要提醒保持专注",
        "Needs Support": "专注度方面需要更多支持",
    },
    "Participation": {
        "Very Active": "课堂参与非常积极",
        "Active": "课堂参与积极",
        "Helpful": "会主动帮助同学",
        "Asks for Help": "遇到问题会主动寻求帮助",
        "Average": "课堂参与度一般",
        "Needs Encouragement": "需要更多鼓励来参与课堂",
    },
    "Practice Speed & Accuracy": {
        "Excellent": "练习速度和准确率都很好",
        "High Accuracy": "练习准确率较高",
        "Fast but Needs Care": "做题速度不错，但需要更仔细",
        "Good but Rushed": "练习表现不错，但有时略急",
        "Good": "练习完成情况不错",
        "Normal": "练习速度正常",
        "Slow but Thoughtful": "做题速度稍慢，但思考比较认真",
        "Needs Support": "练习题方面需要更多支持",
    },
    "Proof Logic": {
        "Strong": "证明逻辑较强",
        "Good with Hint": "在提示下能较好理解证明逻辑",
        "Needs Practice": "证明逻辑还需要更多练习",
    },
    "Calculation": {
        "Good": "计算比较准确",
        "Needs More Care": "计算方面需要更加仔细",
    },
}


OBSERVATION_FIELDS = [
    "Note-taking",
    "Listening & Focus",
    "Participation",
    "Practice Speed & Accuracy",
    "Proof Logic",
    "Calculation",
]


@dataclass
class StudentRow:
    excel_row: int
    values: dict[str, Any]

    @property
    def first_name(self) -> str:
        return str(self.values.get("First Name") or "").strip()

    @property
    def last_name(self) -> str:
        return str(self.values.get("Last Name") or "").strip()

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()

    @property
    def language(self) -> str:
        return str(self.values.get("Parent Language") or "English").strip()

    @property
    def is_absent(self) -> bool:
        return str(self.values.get("Class Attendance") or "").strip().lower() == "absent"


def normalize_uid(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_student_row(workbook_path: Path, sheet_name: str, excel_row: int) -> tuple[Any, StudentRow]:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet {sheet_name!r} not found. Available sheets: {available}")

    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    values = {
        str(header): worksheet.cell(excel_row, col).value
        for col, header in enumerate(headers, start=1)
        if header
    }

    if not values.get("First Name") and not values.get("Last Name"):
        raise ValueError(f"Row {excel_row} does not look like a student row.")

    return workbook, StudentRow(excel_row=excel_row, values=values)


def student_from_worksheet(worksheet: Any, headers: list[Any], excel_row: int) -> StudentRow | None:
    values = {
        str(header): worksheet.cell(excel_row, col).value
        for col, header in enumerate(headers, start=1)
        if header
    }

    if not values.get("First Name") and not values.get("Last Name"):
        return None
    return StudentRow(excel_row=excel_row, values=values)


def find_column(headers: list[Any], column_name: str) -> int:
    try:
        return headers.index(column_name) + 1
    except ValueError as exc:
        raise ValueError(f"The workbook does not have a {column_name!r} column.") from exc


def iter_student_rows(
    worksheet: Any,
    *,
    start_row: int = 2,
    end_row: int | None = None,
) -> list[StudentRow]:
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    final_row = end_row or worksheet.max_row
    students: list[StudentRow] = []
    for row_number in range(start_row, final_row + 1):
        student = student_from_worksheet(worksheet, headers, row_number)
        if student:
            students.append(student)
    return students


def phrase_for(field: str, value: Any, language: str) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text == "Not Observed":
        return None
    custom_phrases = {
        "Chinese": {
            ("Note-taking", "Okay"): "笔记还可以更加完整",
            ("Participation", "Quiet"): "课堂上比较安静",
            ("Practice Speed & Accuracy", "High"): "练习准确率较高",
            ("Practice Speed & Accuracy", "Okay"): "练习完成情况还可以继续提升",
        },
        "English": {
            ("Note-taking", "Okay"): "could make class notes more complete",
            ("Participation", "Quiet"): "was relatively quiet in class",
            ("Practice Speed & Accuracy", "High"): "showed strong accuracy on practice problems",
            ("Practice Speed & Accuracy", "Okay"): "can continue improving practice accuracy",
        },
    }
    custom_language = "Chinese" if language.lower().startswith("chinese") else "English"
    if (phrase := custom_phrases[custom_language].get((field, text))):
        return phrase
    phrases = ZH_PHRASES if language.lower().startswith("chinese") else EN_PHRASES
    return phrases.get(field, {}).get(text, text)


def target_language(student: StudentRow) -> str:
    if student.language.lower().startswith("chinese"):
        return "Chinese"
    return "English"


def revise_remark_with_gpt(student: StudentRow, model: str) -> str:
    remark = str(student.values.get("Remark for Student") or "").strip()
    if not remark:
        return ""

    prompt = f"""
You are revising a teacher's private classroom note before it is saved back to the spreadsheet.

Keep the meaning and all important details.
The teacher note may be informal Chinese. Revise it into smooth, concise Chinese.
Do not add a greeting, parent-facing wording, homework, or information not in the note.
Output only the revised Chinese note.

Student: {student.full_name}
Original note:
{remark}
""".strip()
    return create_response(prompt, model=model)


def additional_comment_for_local_message(student: StudentRow, is_chinese: bool) -> str:
    additional_comment = str(student.values.get(ADDITIONAL_COMMENT_COLUMN) or "").strip()
    if not additional_comment:
        return ""
    if is_chinese:
        return f"另外，{additional_comment}"
    if additional_comment.isascii():
        return f"Additional note: {additional_comment}"
    return ""


def value_from_any_column(student: StudentRow, column_names: tuple[str, ...]) -> str:
    for column_name in column_names:
        value = str(student.values.get(column_name) or "").strip()
        if value:
            return value
    return ""


def personal_feedback_with_gpt(
    student: StudentRow,
    *,
    observations: list[str],
    local_comment: str,
    homework: str | None,
    model: str,
    feedback_type: str = "comprehensive",
) -> str:
    language = target_language(student)
    remark = str(student.values.get("Remark for Student") or "").strip()
    additional_comment = str(student.values.get(ADDITIONAL_COMMENT_COLUMN) or "").strip()
    homework_text = homework or ""
    language_style = (
        ZH_PARENT_COMMENT_STYLE_GUIDE
        if language.lower().startswith("chinese")
        else EN_PARENT_COMMENT_STYLE_GUIDE
    )

    prompt = f"""
You are writing the student-specific part of a parent class update.

Write in {language}.
Follow this teacher style guide:
{PARENT_COMMENT_STYLE_GUIDE}

Language-specific style guide:
{language_style}

Do not mention "not observed".
Do not include attendance.
Do not include the class material paragraph.
Do not end with generic thanks or "thank you for your cooperation".
If Additional Comment is provided, translate it if needed and add it naturally to the end of paragraph 2.
Feedback mode: {feedback_type}.
For general mode, focus on regular classroom feedback and teacher-written notes rather than quiz scores.
For quiz mode, focus on quiz scores, proof-writing issues, and quiz-specific next steps.
For comprehensive mode, include both quiz information and regular classroom feedback.
End with this sentence in Chinese messages: 如果您还有任何问题，可以直接在群里问我，我会尽快回复。
End with this sentence in English messages: If you have any questions, feel free to ask me directly in the group chat. I will reply as soon as possible.

Output paragraph 2 as the personal comment.
If homework feedback is provided, add paragraph 3 as homework feedback.
If there is no homework feedback, output only paragraph 2.

Student: {student.full_name}
Teacher remark, usually in Chinese:
{remark}

Additional Comment:
{additional_comment or "None"}

Structured observations:
{join_naturally(observations, student.language) or "None"}

Local fallback draft:
{local_comment}

Homework feedback:
{homework_text or "None"}
""".strip()
    return create_response(prompt, model=model)


def class_review_paragraph(class_review: str, is_chinese: bool) -> str:
    class_review = class_review.strip()
    if class_review:
        return class_review

    if is_chinese:
        return "今天课堂主要围绕本节几何课的核心概念、例题讲解和课堂练习展开。"
    return (
        "In today's class, we reviewed the main geometry ideas for the lesson, "
        "worked through examples, and practiced applying the methods in class."
    )


def clean_parent_feedback_text(text: str) -> str:
    cleaned = text.replace("：", "，").replace(":", ",")
    cleaned = cleaned.replace("；", "。").replace(";", ".")
    if cleaned.startswith("家长您好"):
        prefix = "家长您好"
        return prefix + cleaned[len(prefix):].replace("家长", "您")
    return cleaned.replace("家长", "您")


def closing_sentence(is_chinese: bool) -> str:
    if is_chinese:
        return "如果您还有任何问题，可以直接在群里问我，我会尽快回复。"
    return "If you have any questions, feel free to ask me directly in the group chat. I will reply as soon as possible."


def append_parent_closing(text: str, is_chinese: bool) -> str:
    closing = closing_sentence(is_chinese)
    cleaned = text.strip()
    if closing in cleaned:
        return cleaned
    return f"{cleaned}\n\n{closing}"


def sentence_join_zh(items: list[str]) -> str:
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    sentences = [items[0]]
    for item in items[1:]:
        if "也" in item or item.startswith(("同时", "特别", "可以", "继续")):
            sentences.append(item)
        elif "还需要" in item:
            sentences.append(item.replace("还需要", "也需要", 1))
        elif "需要" in item:
            sentences.append(item.replace("需要", "也需要", 1))
        else:
            sentences.append(item)
    return "。".join(sentences)


def join_naturally(items: list[str], language: str) -> str:
    if not items:
        return ""
    if language.lower().startswith("chinese"):
        return "，".join(items)
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f", and {items[-1]}"

def homework_paragraph(student: StudentRow, is_chinese: bool) -> str | None:
    homework = str(student.values.get("Homework Reflection") or "").strip()
    if not homework:
        return None
    if is_chinese:
        return f"作业反馈：{homework}"
    return f"Homework feedback: {homework}"

def write_feedback(
    workbook: Any,
    workbook_path: Path,
    sheet_name: str,
    excel_row: int,
    feedback: str,
) -> None:
    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    feedback_col = find_column(headers, "Feedback")

    worksheet.cell(excel_row, feedback_col).value = feedback
    workbook.save(workbook_path)

def write_column_value(
    workbook: Any,
    workbook_path: Path,
    sheet_name: str,
    excel_row: int,
    column_name: str,
    value: str,
) -> None:
    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    column = find_column(headers, column_name)

    worksheet.cell(excel_row, column).value = value
    workbook.save(workbook_path)

def set_column_value(
    worksheet: Any,
    headers: list[Any],
    excel_row: int,
    column_name: str,
    value: str,
) -> None:
    worksheet.cell(excel_row, find_column(headers, column_name)).value = value

