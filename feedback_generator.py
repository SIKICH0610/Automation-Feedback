from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from openai_api import DEFAULT_OPENAI_MODEL, create_response


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = PROJECT_DIR / "Geo_TTh_Student_Script_fixed_rows_only.xlsx"
DEFAULT_SHEET = "Geo TTh"
ADDITIONAL_COMMENT_COLUMN = "Additional Comment"


PARENT_COMMENT_STYLE_GUIDE = """
Write like a real teacher leaving a thoughtful parent update.
Blend quiz performance, classroom behavior, and next steps into natural paragraphs.
Do not use label-style sections, checklist wording, or repeated sentence frames.
Avoid structures like "Regarding the student's performance" or "The main suggestions are".
When several areas need work, vary the phrasing so it sounds spoken rather than copied from a template.
Use natural sentence breaks instead of colons or semicolons.
End with the standard group-chat question sentence.
""".strip()

ZH_PARENT_COMMENT_STYLE_GUIDE = """
中文家长反馈要像老师真实写给家长的消息。
开头可以保留“家长您好”，之后不要反复说“家长”，统一用“您”。
不要写成“关于某某的课堂表现”或“建议后续关注”这种模板句。
如果有多个需要改进的地方，第二个及之后可以自然加入“也”，例如“计算细节也需要更加仔细”。
语气要具体、温和、顺口，像在群里发给家长的说明。
""".strip()

EN_PARENT_COMMENT_STYLE_GUIDE = """
English parent feedback should sound like a natural teacher update.
Do not use report-style labels or repeated template openings.
Connect the quiz result, class habits, and next step in a smooth paragraph.
Keep the tone warm, direct, and parent-friendly.
""".strip()


EN_PHRASES = {
    "Note-taking": {
        "Excellent": "took excellent notes",
        "Good": "took good notes",
        "Needs Reminder": "needed reminders to keep notes complete",
    },
    "Listening & Focus": {
        "Very Focused": "stayed very focused",
        "Focused": "stayed focused",
        "Sometimes Distracted": "was sometimes distracted",
        "Sometimes Needs Reminder": "needed occasional reminders to stay focused",
        "Needs Support": "needed support staying focused",
    },
    "Participation": {
        "Very Active": "participated very actively",
        "Active": "participated actively",
        "Helpful": "helped classmates during practice",
        "Asks for Help": "asked for help when needed",
        "Average": "participated at an average level",
        "Needs Encouragement": "would benefit from more encouragement to participate",
    },
    "Practice Speed & Accuracy": {
        "Excellent": "worked through practice problems with excellent speed and accuracy",
        "High Accuracy": "showed high accuracy on practice problems",
        "Fast but Needs Care": "worked quickly and should continue checking details carefully",
        "Good but Rushed": "did well but sometimes rushed",
        "Good": "did well on practice problems",
        "Normal": "worked at a steady pace",
        "Slow but Thoughtful": "worked slowly but thoughtfully",
        "Needs Support": "needed support with practice problems",
    },
    "Proof Logic": {
        "Strong": "showed strong proof logic",
        "Good with Hint": "understood proof logic well with hints",
        "Needs Practice": "needs more practice with proof logic",
    },
    "Calculation": {
        "Good": "calculated accurately",
        "Needs More Care": "should take more care with calculations",
    },
}


ZH_PHRASES = {
    "Note-taking": {
        "Excellent": "笔记非常完整",
        "Good": "笔记比较完整",
        "Needs Reminder": "笔记方面需要提醒",
    },
    "Listening & Focus": {
        "Very Focused": "听课非常专注",
        "Focused": "听课比较专注",
        "Sometimes Distracted": "有时会分心",
        "Sometimes Needs Reminder": "偶尔需要提醒保持专注",
        "Needs Support": "专注度方面需要更多支持",
    },
    "Participation": {
        "Very Active": "课堂参与非常积极",
        "Active": "课堂参与积极",
        "Helpful": "会主动帮助同学",
        "Asks for Help": "遇到问题会主动寻求帮助",
        "Average": "课堂参与度一般",
        "Needs Encouragement": "需要更多鼓励来参与课堂",
    },
    "Practice Speed & Accuracy": {
        "Excellent": "练习速度和准确率都很好",
        "High Accuracy": "练习准确率较高",
        "Fast but Needs Care": "做题速度不错，但需要更仔细",
        "Good but Rushed": "练习表现不错，但有时略急",
        "Good": "练习完成情况不错",
        "Normal": "练习速度正常",
        "Slow but Thoughtful": "做题速度稍慢，但思考比较认真",
        "Needs Support": "练习题方面需要更多支持",
    },
    "Proof Logic": {
        "Strong": "证明逻辑较强",
        "Good with Hint": "在提示下能较好理解证明逻辑",
        "Needs Practice": "证明逻辑还需要更多练习",
    },
    "Calculation": {
        "Good": "计算比较准确",
        "Needs More Care": "计算方面需要更加仔细",
    },
}


OBSERVATION_FIELDS = [
    "Note-taking",
    "Listening & Focus",
    "Participation",
    "Practice Speed & Accuracy",
    "Proof Logic",
    "Calculation",
]


@dataclass
class StudentRow:
    excel_row: int
    values: dict[str, Any]

    @property
    def first_name(self) -> str:
        return str(self.values.get("First Name") or "").strip()

    @property
    def last_name(self) -> str:
        return str(self.values.get("Last Name") or "").strip()

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()

    @property
    def language(self) -> str:
        return str(self.values.get("Parent Language") or "English").strip()

    @property
    def is_absent(self) -> bool:
        return str(self.values.get("Class Attendance") or "").strip().lower() == "absent"


def normalize_uid(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_student_row(workbook_path: Path, sheet_name: str, excel_row: int) -> tuple[Any, StudentRow]:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet {sheet_name!r} not found. Available sheets: {available}")

    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    values = {
        str(header): worksheet.cell(excel_row, col).value
        for col, header in enumerate(headers, start=1)
        if header
    }

    if not values.get("First Name") and not values.get("Last Name"):
        raise ValueError(f"Row {excel_row} does not look like a student row.")

    return workbook, StudentRow(excel_row=excel_row, values=values)


def student_from_worksheet(worksheet: Any, headers: list[Any], excel_row: int) -> StudentRow | None:
    values = {
        str(header): worksheet.cell(excel_row, col).value
        for col, header in enumerate(headers, start=1)
        if header
    }

    if not values.get("First Name") and not values.get("Last Name"):
        return None
    return StudentRow(excel_row=excel_row, values=values)


def find_column(headers: list[Any], column_name: str) -> int:
    try:
        return headers.index(column_name) + 1
    except ValueError as exc:
        raise ValueError(f"The workbook does not have a {column_name!r} column.") from exc


def iter_student_rows(
    worksheet: Any,
    *,
    start_row: int = 2,
    end_row: int | None = None,
) -> list[StudentRow]:
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    final_row = end_row or worksheet.max_row
    students: list[StudentRow] = []
    for row_number in range(start_row, final_row + 1):
        student = student_from_worksheet(worksheet, headers, row_number)
        if student:
            students.append(student)
    return students


def phrase_for(field: str, value: Any, language: str) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text == "Not Observed":
        return None
    custom_phrases = {
        "Chinese": {
            ("Note-taking", "Okay"): "笔记还可以更加完整",
            ("Participation", "Quiet"): "课堂上比较安静",
            ("Practice Speed & Accuracy", "High"): "练习准确率较高",
            ("Practice Speed & Accuracy", "Okay"): "练习完成情况还可以继续提升",
        },
        "English": {
            ("Note-taking", "Okay"): "could make class notes more complete",
            ("Participation", "Quiet"): "was relatively quiet in class",
            ("Practice Speed & Accuracy", "High"): "showed strong accuracy on practice problems",
            ("Practice Speed & Accuracy", "Okay"): "can continue improving practice accuracy",
        },
    }
    custom_language = "Chinese" if language.lower().startswith("chinese") else "English"
    if (phrase := custom_phrases[custom_language].get((field, text))):
        return phrase
    phrases = ZH_PHRASES if language.lower().startswith("chinese") else EN_PHRASES
    return phrases.get(field, {}).get(text, text)


def target_language(student: StudentRow) -> str:
    if student.language.lower().startswith("chinese"):
        return "Chinese"
    return "English"


def revise_remark_with_gpt(student: StudentRow, model: str) -> str:
    remark = str(student.values.get("Remark for Student") or "").strip()
    if not remark:
        return ""

    prompt = f"""
You are revising a teacher's private classroom note before it is saved back to the spreadsheet.

Keep the meaning and all important details.
The teacher note may be informal Chinese. Revise it into smooth, concise Chinese.
Do not add a greeting, parent-facing wording, homework, or information not in the note.
Output only the revised Chinese note.

Student: {student.full_name}
Original note:
{remark}
""".strip()
    return create_response(prompt, model=model)


def additional_comment_for_local_message(student: StudentRow, is_chinese: bool) -> str:
    additional_comment = str(student.values.get(ADDITIONAL_COMMENT_COLUMN) or "").strip()
    if not additional_comment:
        return ""
    if is_chinese:
        return f"另外，{additional_comment}"
    if additional_comment.isascii():
        return f"Additional note: {additional_comment}"
    return ""


def personal_feedback_with_gpt(
    student: StudentRow,
    *,
    observations: list[str],
    local_comment: str,
    homework: str | None,
    model: str,
) -> str:
    language = target_language(student)
    remark = str(student.values.get("Remark for Student") or "").strip()
    additional_comment = str(student.values.get(ADDITIONAL_COMMENT_COLUMN) or "").strip()
    homework_text = homework or ""
    language_style = (
        ZH_PARENT_COMMENT_STYLE_GUIDE
        if language.lower().startswith("chinese")
        else EN_PARENT_COMMENT_STYLE_GUIDE
    )

    prompt = f"""
You are writing the student-specific part of a parent class update.

Write in {language}.
Follow this teacher style guide:
{PARENT_COMMENT_STYLE_GUIDE}

Language-specific style guide:
{language_style}

Do not mention "not observed".
Do not include attendance.
Do not include the class material paragraph.
Do not end with generic thanks or "thank you for your cooperation".
If Additional Comment is provided, translate it if needed and add it naturally to the end of paragraph 2.
End with this sentence in Chinese messages: 如果您还有任何问题，可以直接在群里问我，我会尽快回复。
End with this sentence in English messages: If you have any questions, feel free to ask me directly in the group chat. I will reply as soon as possible.

Output paragraph 2 as the personal comment.
If homework feedback is provided, add paragraph 3 as homework feedback.
If there is no homework feedback, output only paragraph 2.

Student: {student.full_name}
Teacher remark, usually in Chinese:
{remark}

Additional Comment:
{additional_comment or "None"}

Structured observations:
{join_naturally(observations, student.language) or "None"}

Local fallback draft:
{local_comment}

Homework feedback:
{homework_text or "None"}
""".strip()
    return create_response(prompt, model=model)


def class_review_paragraph(class_review: str, is_chinese: bool) -> str:
    class_review = class_review.strip()
    if class_review:
        return class_review

    if is_chinese:
        return "今天课堂主要围绕本节几何课的核心概念、例题讲解和课堂练习展开。"
    return (
        "In today's class, we reviewed the main geometry ideas for the lesson, "
        "worked through examples, and practiced applying the methods in class."
    )


def clean_parent_feedback_text(text: str) -> str:
    cleaned = text.replace("：", "，").replace(":", ",")
    cleaned = cleaned.replace("；", "。").replace(";", ".")
    if cleaned.startswith("家长您好"):
        prefix = "家长您好"
        return prefix + cleaned[len(prefix):].replace("家长", "您")
    return cleaned.replace("家长", "您")


def closing_sentence(is_chinese: bool) -> str:
    if is_chinese:
        return "如果您还有任何问题，可以直接在群里问我，我会尽快回复。"
    return "If you have any questions, feel free to ask me directly in the group chat. I will reply as soon as possible."


def append_parent_closing(text: str, is_chinese: bool) -> str:
    closing = closing_sentence(is_chinese)
    cleaned = text.strip()
    if closing in cleaned:
        return cleaned
    return f"{cleaned}\n\n{closing}"


def sentence_join_zh(items: list[str]) -> str:
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    sentences = [items[0]]
    for item in items[1:]:
        if "也" in item or item.startswith(("同时", "特别", "可以", "继续")):
            sentences.append(item)
        elif "还需要" in item:
            sentences.append(item.replace("还需要", "也需要", 1))
        elif "需要" in item:
            sentences.append(item.replace("需要", "也需要", 1))
        else:
            sentences.append(item)
    return "。".join(sentences)


def join_naturally(items: list[str], language: str) -> str:
    if not items:
        return ""
    if language.lower().startswith("chinese"):
        return "，".join(items)
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f", and {items[-1]}"


def quiz_score_from_remark(remark: str) -> float | None:
    match = re.search(r"(\d+(?:\.\d+)?)\s*/\s*8", remark)
    if not match:
        return None
    return float(match.group(1))


def format_score(score: float) -> str:
    if score.is_integer():
        return str(int(score))
    return str(score).rstrip("0").rstrip(".")


def strip_quiz_score_text(remark: str) -> str:
    text = re.sub(r"quiz\s*1?\s*[，,]?\s*", "", remark, flags=re.IGNORECASE)
    text = re.sub(r"满分\s*\d+(?:\.\d+)?\s*/\s*8[，,]?", "", text)
    text = re.sub(r"\d+(?:\.\d+)?\s*/\s*8[，,]?", "", text)
    return text.strip(" ，,。.")


def chinese_quiz_score_sentence(score: float) -> str:
    score_text = format_score(score)
    if score < 6.5:
        return (
            f"本次 quiz 得分为 {score_text}/8，低于本次班级平均分 6.9/8，"
            "建议家长最近多关注一下孩子对前半部分知识的掌握情况。"
        )
    if score >= 7.5:
        return f"本次 quiz 得分为 {score_text}/8，整体表现很不错，说明前半部分知识掌握比较扎实。"
    return f"本次 quiz 得分为 {score_text}/8，整体接近或达到本次班级平均水平。"


def english_quiz_score_sentence(score: float) -> str:
    score_text = format_score(score)
    if score < 6.5:
        return (
            f"The quiz score was {score_text}/8, which is below the class average of 6.9/8. "
            "It would be helpful to keep a closer eye on the student's current understanding."
        )
    if score >= 7.5:
        return (
            f"The quiz score was {score_text}/8, which is a strong result and shows solid understanding "
            "of the first-half topics."
        )
    return f"The quiz score was {score_text}/8, which is around the class average."


def chinese_observation_sentence(name: str, observations: list[str], language: str) -> str:
    if not observations:
        return ""
    return f"结合课堂表现来看，{name} {join_naturally(observations, language)}。"


def english_observation_sentence(name: str, observations: list[str], language: str) -> str:
    if not observations:
        return ""
    return f"During class, {name} {join_naturally(observations, language)}."


QUIZ_FOCUS_RULES_ZH = [
    (("选择题全对",), (), "选择题部分完成得很好"),
    (("证明严谨", "proof is rig"), (), "证明书写比较严谨"),
    (("都写了过程",), (), "也能比较完整地写出过程"),
    (("不太爱写过程", "不喜欢写过程"), (), "做题时需要更主动地写出完整过程"),
    (("不仔细", "问题", "错"), ("计算",), "计算细节还需要更加仔细"),
    (("证明不太严谨", "漏步骤"), (), "证明书写需要更严谨，避免漏掉关键步骤"),
    (("先证明三角形全等",), (), "特别要注意先证明三角形全等，再使用对应边或对应角相等"),
    (("reason",), (), "证明中的 reason 需要写得更明确"),
    (("alternate interior angle",), (), "平行线相关证明中要清楚写出 alternate interior angles congruent 等关键理由"),
    (("三种shape", "special angle"), (), "平行线对应的几种 special angle 图形还需要继续巩固，做题时不要被其他平行线干扰判断"),
    (("apple watch", "出怪声"), (), "课堂专注和课堂习惯方面还需要您在家里配合提醒"),
    (("perfect", "8/8"), (), "可以继续保持目前严谨、完整的作答习惯"),
]

QUIZ_FOCUS_RULES_EN = [
    (("perfect", "8/8"), (), "maintaining the strong quiz performance"),
    (("rig", "concise"), ("proof",), "keeping proof writing rigorous and concise"),
    (("calculation", "compute"), (), "checking calculation accuracy"),
    (("reason",), (), "stating proof reasons more clearly"),
    (("process", "steps"), (), "showing complete work and avoiding skipped steps"),
]


def focus_points_from_rules(
    text: str,
    rules: list[tuple[tuple[str, ...], tuple[str, ...], str]],
) -> list[str]:
    lower = text.lower()
    points: list[str] = []
    for any_terms, required_terms, phrase in rules:
        has_any = any(term.lower() in lower for term in any_terms)
        has_required = all(term.lower() in lower for term in required_terms)
        if has_any and has_required:
            points.append(phrase)
    return list(dict.fromkeys(points))


def chinese_quiz_focus_sentences(remark: str) -> list[str]:
    return focus_points_from_rules(remark, QUIZ_FOCUS_RULES_ZH)


def english_quiz_focus_sentences(remark: str) -> list[str]:
    return focus_points_from_rules(remark, QUIZ_FOCUS_RULES_EN)


def quiz_comment_paragraph(student: StudentRow, observations: list[str], is_chinese: bool) -> str | None:
    remark = str(student.values.get("Remark for Student") or "").strip()
    if "quiz" not in remark.lower() and "/8" not in remark:
        return None

    name = student.first_name or student.full_name
    score = quiz_score_from_remark(remark)
    additional_comment = additional_comment_for_local_message(student, is_chinese)

    if is_chinese:
        sentences: list[str] = []
        if score is not None:
            sentences.append(chinese_quiz_score_sentence(score))

        focus = chinese_quiz_focus_sentences(remark)
        if focus:
            focus_text = sentence_join_zh(focus)
            if score is not None and score >= 7.5:
                sentences.append(f"{name} 这次表现比较扎实，后续可以继续保持，尤其是{focus_text}。")
            else:
                sentences.append(f"从这次 quiz 和课堂情况来看，{name} 接下来可以重点关注{focus_text}。")
        else:
            cleaned = strip_quiz_score_text(remark)
            if cleaned:
                sentences.append(f"从本次课堂记录来看，{name}{cleaned}。")

        if observations:
            sentences.append(chinese_observation_sentence(name, observations, student.language))
        if additional_comment:
            sentences.append(additional_comment + "。")
        sentences.append("后续可以继续复习前半部分知识，并在计算准确性和证明步骤完整性上多检查。")
        return "".join(sentences)

    sentences = []
    if score is not None:
        sentences.append(english_quiz_score_sentence(score))

    focus = english_quiz_focus_sentences(remark)
    if focus:
        focus_text = ", and ".join(focus)
        sentences.append(f"{name} can build from this by continuing to focus on {focus_text}.")
    elif remark.isascii():
        cleaned = strip_quiz_score_text(remark)
        if cleaned:
            sentences.append(f"My classroom note for {name} is that {cleaned}.")

    if observations:
        sentences.append(english_observation_sentence(name, observations, student.language))
    if additional_comment:
        sentences.append(additional_comment + ".")
    sentences.append("Going forward, reviewing the first-half topics while checking calculation details and proof structure carefully will help keep the foundation strong.")
    return " ".join(sentences)


def comment_paragraph(student: StudentRow, observations: list[str], is_chinese: bool) -> str:
    name = student.first_name or student.full_name
    remark = str(student.values.get("Remark for Student") or "").strip()
    additional_comment = additional_comment_for_local_message(student, is_chinese)

    quiz_comment = quiz_comment_paragraph(student, observations, is_chinese)
    if quiz_comment:
        return quiz_comment

    if is_chinese:
        if remark:
            base = f"{name}今天课堂中{remark}。"
            if additional_comment:
                return f"{base}{additional_comment}。"
            return base
        if observations:
            base = (
                f"{name}今天整体表现稳定，"
                f"{join_naturally(observations, student.language)}。之后可以继续保持好的课堂习惯，"
                "同时在证明逻辑和细节检查上多练习。"
            )
            if additional_comment:
                return f"{base}{additional_comment}。"
            return base
        return f"{name}今天的课堂表现已记录，之后可以继续保持稳定的学习节奏。"

    if observations:
        base = (
            f"{name} had a steady class today and "
            f"{join_naturally(observations, student.language)}. Continuing to build proof logic "
            "while checking details carefully will be helpful."
        )
        if additional_comment:
            return f"{base} {additional_comment}."
        return base
    if remark and remark.isascii():
        base = f"My classroom note for {name} is that {remark}"
        if additional_comment:
            return f"{base} {additional_comment}."
        return base
    return f"{name}'s classroom notes have been recorded for today's lesson."


def homework_paragraph(student: StudentRow, is_chinese: bool) -> str | None:
    homework = str(student.values.get("Homework Reflection") or "").strip()
    if not homework:
        return None
    if is_chinese:
        return f"作业反馈：{homework}"
    return f"Homework feedback: {homework}"


def generate_feedback(
    student: StudentRow,
    class_review: str = "",
    *,
    use_api: bool = False,
    model: str = DEFAULT_OPENAI_MODEL,
) -> str | None:
    if student.is_absent:
        return None

    language = student.language
    is_chinese = language.lower().startswith("chinese")
    observations = [
        phrase
        for field in OBSERVATION_FIELDS
        if (phrase := phrase_for(field, student.values.get(field), language))
    ]
    class_paragraph = class_review_paragraph(class_review, is_chinese)
    local_comment = comment_paragraph(student, observations, is_chinese)
    homework = homework_paragraph(student, is_chinese)

    if use_api:
        personal_section = personal_feedback_with_gpt(
            student,
            observations=observations,
            local_comment=local_comment,
            homework=homework,
            model=model,
        )
        feedback = clean_parent_feedback_text("\n\n".join([class_paragraph, personal_section.strip()]))
        return append_parent_closing(feedback, is_chinese)

    paragraphs = [class_paragraph, local_comment]
    if homework:
        paragraphs.append(homework)
    feedback = clean_parent_feedback_text("\n\n".join(paragraphs))
    return append_parent_closing(feedback, is_chinese)


def write_feedback(
    workbook: Any,
    workbook_path: Path,
    sheet_name: str,
    excel_row: int,
    feedback: str,
) -> None:
    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    feedback_col = find_column(headers, "Feedback")

    worksheet.cell(excel_row, feedback_col).value = feedback
    workbook.save(workbook_path)


def write_column_value(
    workbook: Any,
    workbook_path: Path,
    sheet_name: str,
    excel_row: int,
    column_name: str,
    value: str,
) -> None:
    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    column = find_column(headers, column_name)

    worksheet.cell(excel_row, column).value = value
    workbook.save(workbook_path)


def set_column_value(
    worksheet: Any,
    headers: list[Any],
    excel_row: int,
    column_name: str,
    value: str,
) -> None:
    worksheet.cell(excel_row, find_column(headers, column_name)).value = value


def print_student_result(student: StudentRow, feedback: str | None) -> None:
    print("=" * 72)
    print(f"Row: {student.excel_row}")
    print(f"Student: {student.full_name}")
    print(f"UID: {normalize_uid(student.values.get('uid'))}")
    print()
    if feedback is None:
        print("Skipped: student was absent, so no feedback comment was generated.")
    else:
        print(feedback)
    print()


def export_review_csv(rows: list[dict[str, str]], output_path: Path) -> None:
    fieldnames = [
        "row",
        "uid",
        "student",
        "status",
        "feedback_column",
        "revised_remark",
        "feedback",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate one parent-facing student feedback entry from the Excel tracker."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument(
        "--row",
        type=int,
        default=2,
        help="Excel row number to generate. Row 2 is the first student row.",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Generate feedback for every student row in the selected sheet.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="First row to use with --all. Defaults to 2.",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        help="Last row to use with --all. Omit to continue through the sheet.",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="Write generated text back to the selected feedback column. Preview-only by default.",
    )
    parser.add_argument(
        "--feedback-column",
        default="Feedback",
        help=(
            "Workbook column to write generated comments into. Examples: "
            "'Pre-quiz informing', 'Quiz Feedback', 'Second Feeback', or 'Feedback'."
        ),
    )
    parser.add_argument(
        "--review-csv",
        type=Path,
        help="Save generated previews to a CSV for review before or while writing to Excel.",
    )
    parser.add_argument(
        "--class-review",
        default="",
        help="What the class covered today. This becomes the first paragraph.",
    )
    parser.add_argument(
        "--class-review-file",
        type=Path,
        help="Optional text file containing the class review paragraph.",
    )
    parser.add_argument(
        "--class-review-file-zh",
        type=Path,
        help="Optional Chinese class review file for Chinese parent messages.",
    )
    parser.add_argument(
        "--class-review-file-en",
        type=Path,
        help="Optional English class review file for English parent messages.",
    )
    parser.add_argument(
        "--use-api",
        action="store_true",
        help="Use the OpenAI API to polish the parent-facing personal comment.",
    )
    parser.add_argument("--model", default=DEFAULT_OPENAI_MODEL)
    parser.add_argument(
        "--revise-remark",
        action="store_true",
        help="Use the OpenAI API to revise the Remark for Student cell text first.",
    )
    parser.add_argument(
        "--write-revised-remark",
        action="store_true",
        help="Save the revised Remark for Student text back to the workbook.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    class_review = args.class_review
    if args.class_review_file:
        class_review = args.class_review_file.read_text(encoding="utf-8").strip()
    class_review_zh = (
        args.class_review_file_zh.read_text(encoding="utf-8").strip()
        if args.class_review_file_zh
        else class_review
    )
    class_review_en = (
        args.class_review_file_en.read_text(encoding="utf-8").strip()
        if args.class_review_file_en
        else class_review
    )

    workbook = load_workbook(args.workbook)
    if args.sheet not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet {args.sheet!r} not found. Available sheets: {available}")

    worksheet = workbook[args.sheet]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    if args.write:
        find_column(headers, args.feedback_column)
    if args.all:
        students = iter_student_rows(
            worksheet,
            start_row=args.start_row,
            end_row=args.end_row,
        )
    else:
        student = student_from_worksheet(worksheet, headers, args.row)
        if not student:
            raise ValueError(f"Row {args.row} does not look like a student row.")
        students = [student]

    print(f"Sheet: {args.sheet}")
    print(f"Mode: {'all rows' if args.all else 'single row'}")
    print(f"Write feedback: {'yes' if args.write else 'no, preview only'}")
    print(f"Feedback column: {args.feedback_column}")
    print()

    generated = 0
    skipped = 0
    review_rows: list[dict[str, str]] = []
    for student in students:
        revised_remark = ""
        if args.revise_remark:
            revised_remark = revise_remark_with_gpt(student, args.model)
            if revised_remark:
                student.values["Remark for Student"] = revised_remark
                print("=" * 72)
                print(f"Row {student.excel_row} revised remark:")
                print(revised_remark)
                print()
                if args.write_revised_remark:
                    set_column_value(
                        worksheet,
                        headers,
                        student.excel_row,
                        "Remark for Student",
                        revised_remark,
                    )

        student_class_review = (
            class_review_zh if student.language.lower().startswith("chinese") else class_review_en
        )

        feedback = generate_feedback(
            student,
            class_review=student_class_review,
            use_api=args.use_api,
            model=args.model,
        )
        print_student_result(student, feedback)

        review_rows.append(
            {
                "row": str(student.excel_row),
                "uid": normalize_uid(student.values.get("uid")),
                "student": student.full_name,
                "status": "skipped_absent" if feedback is None else "generated",
                "feedback_column": args.feedback_column,
                "revised_remark": revised_remark,
                "feedback": feedback or "",
            }
        )

        if feedback is None:
            skipped += 1
            continue

        generated += 1
        if args.write:
            set_column_value(worksheet, headers, student.excel_row, args.feedback_column, feedback)

    if args.write or args.write_revised_remark:
        workbook.save(args.workbook)
        print(f"Saved workbook: {args.workbook}")

    if args.review_csv:
        export_review_csv(review_rows, args.review_csv)
        print(f"Saved review CSV: {args.review_csv}")

    print(f"Done. Generated: {generated}. Skipped: {skipped}.")


if __name__ == "__main__":
    main()
