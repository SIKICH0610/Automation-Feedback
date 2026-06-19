from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from pathlib import Path
import shutil
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_TARGET = PROJECT_DIR / "Geo_TTh_Student_Script_fixed_rows_only.xlsx"
DEFAULT_SHEETS = ["AMC10 Intro", "Geo WF", "Geo TTh", "Geo MTWThF"]
OBSERVATION_FIELDS = {
    "Note-taking",
    "Listening & Focus",
    "Participation",
    "Practice Speed & Accuracy",
    "Proof Logic",
    "Calculation",
}


def header_values(worksheet: Any) -> list[str]:
    return [
        str(worksheet.cell(1, col).value).strip()
        for col in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, col).value
    ]


def row_values_by_header(worksheet: Any, headers: list[str], row_number: int) -> dict[str, Any]:
    return {
        header: worksheet.cell(row_number, col).value
        for col, header in enumerate(headers, start=1)
    }


def normalize_uid(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_value(header: str, value: Any, existing_value: Any = None) -> Any:
    if header == "uid":
        return normalize_uid(value)

    if header == "Class Attendance":
        if value is True:
            return "Present"
        if value is False:
            return "Absent"
        text = str(value or "").strip().lower()
        if text == "true":
            return "Present"
        if text == "false":
            return "Absent"
        return value

    if header in OBSERVATION_FIELDS and isinstance(value, bool):
        if existing_value not in (None, "", True, False):
            return existing_value
        return None

    return value


def copy_header_template(source_sheet: Any, target_sheet: Any) -> None:
    for col in range(1, source_sheet.max_column + 1):
        source_cell = source_sheet.cell(1, col)
        target_cell = target_sheet.cell(1, col)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        target_sheet.column_dimensions[target_cell.column_letter].width = (
            source_sheet.column_dimensions[source_cell.column_letter].width
        )


def ensure_target_sheet(target_workbook: Any, sheet_name: str) -> Any:
    if sheet_name in target_workbook.sheetnames:
        return target_workbook[sheet_name]

    template = target_workbook["Geo TTh"] if "Geo TTh" in target_workbook.sheetnames else target_workbook.active
    new_sheet = target_workbook.create_sheet(sheet_name)
    copy_header_template(template, new_sheet)
    return new_sheet


def existing_rows_by_uid(target_sheet: Any, headers: list[str]) -> dict[Any, dict[str, Any]]:
    rows: dict[Any, dict[str, Any]] = {}
    for row_number in range(2, target_sheet.max_row + 1):
        row = row_values_by_header(target_sheet, headers, row_number)
        uid = normalize_uid(row.get("uid"))
        if uid:
            rows[uid] = row
    return rows


def clear_data_rows(target_sheet: Any, max_columns: int) -> None:
    for row in target_sheet.iter_rows(min_row=2, max_row=target_sheet.max_row, max_col=max_columns):
        for cell in row:
            cell.value = None


def nonempty_source_rows(source_sheet: Any) -> list[int]:
    row_numbers: list[int] = []
    for row_number in range(2, source_sheet.max_row + 1):
        values = [source_sheet.cell(row_number, col).value for col in range(1, source_sheet.max_column + 1)]
        if any(value not in (None, "") for value in values):
            row_numbers.append(row_number)
    return row_numbers


def load_sheet(source_workbook: Any, target_workbook: Any, sheet_name: str) -> int:
    source_sheet = source_workbook[sheet_name]
    target_sheet = ensure_target_sheet(target_workbook, sheet_name)

    source_headers = header_values(source_sheet)
    target_headers = header_values(target_sheet)
    existing_by_uid = existing_rows_by_uid(target_sheet, target_headers)
    clear_data_rows(target_sheet, len(target_headers))

    written = 0
    for target_row, source_row_number in enumerate(nonempty_source_rows(source_sheet), start=2):
        source_row = row_values_by_header(source_sheet, source_headers, source_row_number)
        uid = normalize_uid(source_row.get("uid"))
        existing_row = existing_by_uid.get(uid, {})

        for col, header in enumerate(target_headers, start=1):
            if header in source_row:
                value = normalize_value(header, source_row.get(header), existing_row.get(header))
            else:
                value = existing_row.get(header)
            target_sheet.cell(target_row, col).value = value
        written += 1

    return written


def main() -> None:
    parser = argparse.ArgumentParser(description="Load selected class roster entries into the local feedback workbook.")
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--target", default=DEFAULT_TARGET, type=Path)
    parser.add_argument("--sheets", nargs="+", default=DEFAULT_SHEETS)
    args = parser.parse_args()

    source_workbook = load_workbook(args.source, data_only=False)
    target_workbook = load_workbook(args.target)

    missing = [sheet for sheet in args.sheets if sheet not in source_workbook.sheetnames]
    if missing:
        raise ValueError(f"Source workbook is missing requested sheets: {', '.join(missing)}")

    backup_dir = PROJECT_DIR / "outputs" / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{args.target.stem}.backup_{timestamp}{args.target.suffix}"
    shutil.copy2(args.target, backup_path)

    for sheet_name in args.sheets:
        written = load_sheet(source_workbook, target_workbook, sheet_name)
        print(f"{sheet_name}: loaded {written} entries")

    target_workbook.save(args.target)
    print(f"Saved: {args.target}")
    print(f"Backup: {backup_path}")


if __name__ == "__main__":
    main()
