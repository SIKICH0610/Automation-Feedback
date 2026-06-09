from __future__ import annotations

import argparse
from datetime import datetime
from dataclasses import dataclass
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import time
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook

from feedback_generator import (
    DEFAULT_SHEET,
    DEFAULT_WORKBOOK,
    StudentRow,
    generate_feedback,
    normalize_uid,
    student_from_worksheet,
)
from openai_api import DEFAULT_OPENAI_MODEL


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


DEFAULT_WECOM_TITLE_RE = r".*(WeCom|企业微信|WXWork).*"
DEFAULT_WHATSAPP_TITLE_RE = r".*(WhatsApp|WhatsApp Web).*"
STATUS_COLUMN = "Send Status"
ERROR_COLUMN = "Send Error"
LAST_ATTEMPT_COLUMN = "Last Attempt"
PREFERRED_CHANNEL_COLUMN = "Preferred Channel"
WHATSAPP_PHONE_COLUMN = "WhatsApp Phone"
WHATSAPP_SEARCH_KEY_COLUMN = "WhatsApp Search Key"
WHATSAPP_TARGET_TYPE_COLUMN = "WhatsApp Target Type"


@dataclass
class DesktopAppSpec:
    key: str
    display_name: str
    title_re: str
    process_names: tuple[str, ...]
    env_path_var: str
    exe_names: tuple[str, ...]
    common_paths: tuple[str, ...]
    uri: str | None = None


@dataclass
class DesktopAppStatus:
    key: str
    display_name: str
    dependency_ok: bool
    process_running: bool
    window_found: bool
    launch_attempted: bool = False
    launched: bool = False
    message: str = ""
    dependency_error: str = ""


@dataclass
class PasteJob:
    excel_row: int
    uid: str
    student_name: str
    parent_language: str
    channel: str
    search_key: str
    expected_chat_name: str
    whatsapp_phone: str
    whatsapp_target_type: str
    feedback: str


@dataclass
class JobResult:
    status: str
    pasted: bool = False
    error: str = ""


@dataclass
class JobLoadError:
    row_number: int
    status: str
    error: str


def build_app_specs(
    *,
    wecom_title_re: str,
    whatsapp_title_re: str,
) -> dict[str, DesktopAppSpec]:
    local_app_data = os.environ.get("LOCALAPPDATA", "")
    program_files = os.environ.get("ProgramFiles", r"C:\Program Files")
    program_files_x86 = os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")

    return {
        "wecom": DesktopAppSpec(
            key="wecom",
            display_name="WeCom / 企业微信",
            title_re=wecom_title_re,
            process_names=("WXWork.exe", "WeCom.exe"),
            env_path_var="WECOM_EXE",
            exe_names=("WXWork.exe", "WeCom.exe"),
            common_paths=(
                rf"{program_files_x86}\Tencent\WeCom\WXWork.exe",
                rf"{program_files}\Tencent\WeCom\WXWork.exe",
                rf"{local_app_data}\Tencent\WXWork\WXWork.exe",
                rf"{local_app_data}\WXWork\WXWork.exe",
            ),
        ),
        "whatsapp": DesktopAppSpec(
            key="whatsapp",
            display_name="WhatsApp / WhatsApp Web",
            title_re=whatsapp_title_re,
            process_names=("WhatsApp.exe", "chrome.exe", "msedge.exe", "brave.exe", "firefox.exe"),
            env_path_var="WHATSAPP_EXE",
            exe_names=("WhatsApp.exe",),
            common_paths=(
                rf"{local_app_data}\WhatsApp\WhatsApp.exe",
                rf"{program_files}\WindowsApps\WhatsApp.exe",
            ),
            uri="https://web.whatsapp.com/",
        ),
    }


def header_values(worksheet: Any) -> list[Any]:
    return [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]


def automation_dependency_status() -> tuple[bool, str]:
    try:
        import pywinauto  # noqa: F401
        import pyperclip  # noqa: F401
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"
    return True, ""


def process_is_running(process_names: tuple[str, ...]) -> bool:
    try:
        output = subprocess.check_output(
            ["tasklist"],
            text=True,
            encoding="utf-8",
            errors="ignore",
        )
    except Exception:
        return False

    output_lower = output.lower()
    return any(process_name.lower() in output_lower for process_name in process_names)


def app_window_found(title_re: str) -> bool:
    return find_window_by_title_re(title_re) is not None


def find_window_by_title_re(title_re: str) -> Any | None:
    try:
        from pywinauto import Desktop
    except Exception:
        return None

    desktop = Desktop(backend="uia")
    pattern = re.compile(title_re)
    try:
        for window in desktop.windows():
            try:
                title = window.window_text().strip()
            except Exception:
                continue
            if title and pattern.search(title):
                return window
    except Exception:
        pass

    try:
        window = desktop.window(title_re=title_re)
        if window.exists(timeout=1):
            return window
    except Exception:
        return None
    return None


def app_status(spec: DesktopAppSpec) -> DesktopAppStatus:
    dependency_ok, dependency_error = automation_dependency_status()
    running = process_is_running(spec.process_names)
    window_found = app_window_found(spec.title_re) if dependency_ok else False

    if not dependency_ok:
        message = "desktop automation dependencies are not available"
    elif window_found:
        message = "window found"
    elif running:
        message = "process is running, but matching window was not found"
    else:
        message = "not running"

    return DesktopAppStatus(
        key=spec.key,
        display_name=spec.display_name,
        dependency_ok=dependency_ok,
        process_running=running,
        window_found=window_found,
        message=message,
        dependency_error=dependency_error,
    )


def launch_candidates(spec: DesktopAppSpec, configured_path: Path | None) -> list[str]:
    candidates: list[str] = []
    if configured_path:
        candidates.append(str(configured_path))

    env_path = os.environ.get(spec.env_path_var, "").strip()
    if env_path:
        candidates.append(env_path)

    for exe_name in spec.exe_names:
        found = shutil.which(exe_name)
        if found:
            candidates.append(found)

    candidates.extend(spec.common_paths)
    return candidates


def launch_app(spec: DesktopAppSpec, configured_path: Path | None = None) -> bool:
    for candidate in launch_candidates(spec, configured_path):
        candidate_path = Path(candidate).expanduser()
        if not candidate_path.exists():
            continue
        subprocess.Popen([str(candidate_path)])
        return True

    if spec.uri:
        try:
            os.startfile(spec.uri)  # type: ignore[attr-defined]
            return True
        except OSError:
            return False

    return False


def ensure_app_available(
    spec: DesktopAppSpec,
    *,
    configured_path: Path | None = None,
    open_if_missing: bool = False,
    settle_seconds: float = 3,
) -> DesktopAppStatus:
    status = app_status(spec)
    if status.window_found or not open_if_missing:
        return status
    if not status.dependency_ok:
        return status

    launched = launch_app(spec, configured_path)
    status.launch_attempted = True
    status.launched = launched
    if not launched:
        status.message = (
            f"could not launch automatically; open {spec.display_name} manually or set "
            f"{spec.env_path_var} / pass an explicit exe path"
        )
        return status

    time.sleep(settle_seconds)
    refreshed = app_status(spec)
    refreshed.launch_attempted = True
    refreshed.launched = launched
    if not refreshed.window_found:
        refreshed.message = "launch attempted, but matching window was not found yet"
    return refreshed


def print_app_status(status: DesktopAppStatus) -> None:
    print(f"{status.display_name}: {status.message}")
    print(f"  dependency ok: {'yes' if status.dependency_ok else 'no'}")
    if status.dependency_error:
        print(f"  dependency error: {status.dependency_error}")
        print("  install command: py -3.11 -m pip install -r requirements.txt")
    print(f"  process running: {'yes' if status.process_running else 'no'}")
    print(f"  window found: {'yes' if status.window_found else 'no'}")
    if status.launch_attempted:
        print(f"  launch attempted: yes")
        print(f"  launch command accepted: {'yes' if status.launched else 'no'}")


def print_matching_window_titles(title_re: str) -> None:
    try:
        from pywinauto import Desktop
    except Exception as exc:
        print(f"Could not inspect windows: {type(exc).__name__}: {exc}")
        return

    pattern = re.compile(title_re)
    found = False
    for window in Desktop(backend="uia").windows():
        try:
            title = window.window_text().strip()
        except Exception:
            continue
        if title and pattern.search(title):
            found = True
            print(f"Matched window title: {title}")

    if not found:
        print(f"No top-level windows matched: {title_re}")


def value_for(student: StudentRow, column_name: str) -> str:
    return str(student.values.get(column_name) or "").strip()


def normalize_channel(value: str) -> str:
    text = value.strip().lower()
    if text in {"wechat", "wecom", "wxwork", "企业微信"}:
        return "wecom"
    if text in {"whatsapp", "whatapp", "wa"}:
        return "whatsapp"
    return ""


def choose_channel(student: StudentRow) -> str:
    configured_channel = normalize_channel(value_for(student, PREFERRED_CHANNEL_COLUMN))
    if configured_channel:
        return configured_channel

    language = value_for(student, "Parent Language").lower()
    if language.startswith("chinese"):
        return "wecom"
    return "whatsapp"


def normalize_whatsapp_phone(phone: str) -> str:
    return "".join(character for character in phone if character.isdigit())


def whatsapp_target_type(student: StudentRow) -> str:
    configured_type = value_for(student, WHATSAPP_TARGET_TYPE_COLUMN).strip().lower()
    if configured_type in {"phone", "direct", "direct_phone"}:
        return "phone"
    if configured_type in {"group", "group_search", "search"}:
        return "group_search"
    if value_for(student, WHATSAPP_PHONE_COLUMN):
        return "phone"
    return "group_search"


def build_expected_chat_name(student: StudentRow, uid: str) -> str:
    configured_name = value_for(student, "WeCom Group Name")
    if configured_name:
        return configured_name
    return f"{student.full_name} - {uid}".strip()


def build_search_key(student: StudentRow, channel: str, uid: str) -> str:
    if channel == "whatsapp":
        return value_for(student, WHATSAPP_SEARCH_KEY_COLUMN) or uid
    return uid


def clear_clipboard() -> None:
    try:
        import pyperclip

        pyperclip.copy("")
    except Exception:
        pass


def timestamp() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_status_columns(workbook_path: Path, sheet_name: str) -> None:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return

    worksheet = workbook[sheet_name]
    headers = header_values(worksheet)
    changed = False
    for column_name in (STATUS_COLUMN, ERROR_COLUMN, LAST_ATTEMPT_COLUMN):
        if column_name not in headers:
            worksheet.cell(1, worksheet.max_column + 1).value = column_name
            headers.append(column_name)
            changed = True

    if changed:
        workbook.save(workbook_path)
    workbook.close()


def set_status_value(worksheet: Any, headers: list[Any], row_number: int, column_name: str, value: str) -> None:
    if column_name not in headers:
        worksheet.cell(1, len(headers) + 1).value = column_name
        headers.append(column_name)
    worksheet.cell(row_number, headers.index(column_name) + 1).value = value


def write_job_status(
    workbook_path: Path,
    sheet_name: str,
    row_number: int,
    *,
    status: str,
    error: str = "",
) -> None:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheet_name]
    headers = header_values(worksheet)
    set_status_value(worksheet, headers, row_number, STATUS_COLUMN, status)
    set_status_value(worksheet, headers, row_number, ERROR_COLUMN, error)
    set_status_value(worksheet, headers, row_number, LAST_ATTEMPT_COLUMN, timestamp())
    workbook.save(workbook_path)
    workbook.close()


def feedback_for_student(
    student: StudentRow,
    *,
    class_review: str,
    use_api: bool,
    model: str,
) -> str:
    existing_feedback = value_for(student, "Feedback")
    if existing_feedback:
        return existing_feedback

    feedback = generate_feedback(
        student,
        class_review=class_review,
        use_api=use_api,
        model=model,
    )
    if not feedback:
        raise ValueError(f"Row {student.excel_row} has no feedback because the student is absent.")
    return feedback


def build_paste_job(
    student: StudentRow,
    *,
    class_review: str,
    use_api: bool,
    model: str,
) -> PasteJob:
    uid = normalize_uid(student.values.get("uid"))
    if not uid:
        raise ValueError(f"Row {student.excel_row} does not have a uid.")

    channel = choose_channel(student)
    phone = normalize_whatsapp_phone(value_for(student, WHATSAPP_PHONE_COLUMN))
    target_type = whatsapp_target_type(student)

    group_chat_enabled = value_for(student, "Group Chat").lower()
    if channel == "wecom" and group_chat_enabled in {"false", "no", "0"}:
        raise ValueError(f"Row {student.excel_row} is not marked as using a group chat.")

    if channel == "whatsapp" and target_type == "phone" and not phone:
        raise ValueError(
            f"Row {student.excel_row} uses WhatsApp phone mode but does not have {WHATSAPP_PHONE_COLUMN!r}."
        )

    return PasteJob(
        excel_row=student.excel_row,
        uid=uid,
        student_name=student.full_name,
        parent_language=value_for(student, "Parent Language") or "English",
        channel=channel,
        search_key=build_search_key(student, channel, uid),
        expected_chat_name=build_expected_chat_name(student, uid),
        whatsapp_phone=phone,
        whatsapp_target_type=target_type,
        feedback=feedback_for_student(
            student,
            class_review=class_review,
            use_api=use_api,
            model=model,
        ),
    )


def parse_row_numbers(row_spec: str) -> list[int]:
    rows: list[int] = []
    for part in row_spec.split(","):
        item = part.strip()
        if not item:
            continue
        if "-" in item:
            start_text, end_text = item.split("-", 1)
            start = int(start_text.strip())
            end = int(end_text.strip())
            if end < start:
                raise ValueError(f"Invalid row range {item!r}: end row is before start row.")
            rows.extend(range(start, end + 1))
        else:
            rows.append(int(item))

    if not rows:
        raise ValueError("--rows did not contain any row numbers.")
    return rows


def selected_row_numbers(args: argparse.Namespace) -> list[int]:
    if args.rows:
        return parse_row_numbers(args.rows)

    if args.start_row is not None or args.end_row is not None:
        start = args.start_row if args.start_row is not None else args.row
        end = args.end_row if args.end_row is not None else start
        if end < start:
            raise ValueError("--end-row cannot be before --start-row.")
        return list(range(start, end + 1))

    return [args.row]


def load_jobs(
    workbook_path: Path,
    *,
    sheet_name: str,
    row_numbers: list[int],
    class_review: str,
    use_api: bool,
    model: str,
) -> list[PasteJob | JobLoadError]:
    workbook = load_workbook(workbook_path, read_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet {sheet_name!r} not found. Available sheets: {available}")

    worksheet = workbook[sheet_name]
    headers = header_values(worksheet)
    jobs: list[PasteJob | JobLoadError] = []
    for row_number in row_numbers:
        student = student_from_worksheet(worksheet, headers, row_number)
        if not student:
            jobs.append(
                JobLoadError(
                    row_number=row_number,
                    status="needs_review",
                    error=f"Row {row_number} does not look like a student row.",
                )
            )
            continue

        try:
            jobs.append(
                build_paste_job(
                    student,
                    class_review=class_review,
                    use_api=use_api,
                    model=model,
                )
            )
        except ValueError as exc:
            status = "skipped_absent" if "absent" in str(exc).lower() else "needs_review"
            jobs.append(JobLoadError(row_number=row_number, status=status, error=str(exc)))

    workbook.close()
    return jobs


class WeComPasteRobot:
    def __init__(
        self,
        *,
        title_re: str = DEFAULT_WECOM_TITLE_RE,
        search_shortcut: str = "^f",
        settle_seconds: float = 0.8,
    ) -> None:
        try:
            from pywinauto import Desktop
            from pywinauto.keyboard import send_keys
        except ImportError as exc:
            raise RuntimeError(
                "Install desktop automation dependencies first: "
                "python -m pip install pywinauto pyperclip"
            ) from exc

        self.desktop = Desktop(backend="uia")
        self.send_keys = send_keys
        self.title_re = title_re
        self.search_shortcut = search_shortcut
        self.settle_seconds = settle_seconds
        self.window: Any | None = None

    def focus_window(self) -> Any:
        if self.window is not None:
            try:
                if self.window.exists(timeout=0.5):
                    self.window.set_focus()
                    return self.window
            except Exception:
                self.window = None

        window = self.desktop.window(title_re=self.title_re)
        window.set_focus()
        self.window = window
        return window

    def visible_text(self) -> str:
        window = self.focus_window()
        texts: list[str] = []
        try:
            title = window.window_text().strip()
        except Exception:
            title = ""
        if title:
            texts.append(title)

        try:
            import win32gui

            foreground_title = win32gui.GetWindowText(win32gui.GetForegroundWindow()).strip()
        except Exception:
            foreground_title = ""
        if foreground_title and foreground_title not in texts:
            texts.append(f"foreground window: {foreground_title}")

        for element in window.descendants():
            try:
                text = element.window_text().strip()
            except Exception:
                continue
            if text:
                texts.append(text)
        return "\n".join(texts)

    def search_chat(self, search_key: str) -> None:
        try:
            import pyperclip
        except ImportError as exc:
            raise RuntimeError("Install clipboard dependency first: python -m pip install pyperclip") from exc

        self.focus_window()
        self.send_keys(self.search_shortcut)
        time.sleep(0.2)
        pyperclip.copy(search_key)
        self.send_keys("^v")
        time.sleep(self.settle_seconds)

    def search_result_candidates(self, job: PasteJob) -> list[tuple[int, Any, str]]:
        window = self.focus_window()
        blocked_phrases = (
            "search for mobile",
            "global search",
            "ctrl+alt+f",
        )
        name_parts = [part for part in job.student_name.split() if part]
        elements: list[tuple[Any, str, Any, str]] = []
        blocked_tops: list[int] = []
        candidates: list[tuple[int, Any, str]] = []

        for element in window.descendants():
            try:
                text = element.window_text().strip()
                rectangle = element.rectangle()
            except Exception:
                continue
            if not text or text == job.search_key:
                continue

            lowered = text.lower()
            if any(phrase in lowered for phrase in blocked_phrases):
                blocked_tops.append(rectangle.top)

            elements.append((element, text, rectangle, lowered))

        first_blocked_top = min(blocked_tops) if blocked_tops else None

        for element, text, rectangle, lowered in elements:
            if any(phrase in lowered for phrase in blocked_phrases):
                continue
            if first_blocked_top is not None and rectangle.top >= first_blocked_top - 2:
                continue

            score = 0
            if job.uid in text:
                score += 100
            if job.expected_chat_name and job.expected_chat_name in text:
                score += 60
            if any(part in text for part in name_parts):
                score += 30
            if "external" in lowered:
                score += 10
            if "group chat name" in lowered:
                score -= 25

            if score > 0 and rectangle.width() > 20 and rectangle.height() > 8:
                candidates.append((score, element, text))

        candidates.sort(key=lambda candidate: candidate[0], reverse=True)
        return candidates

    def search_overlay_visible(self) -> bool:
        text = self.visible_text().lower()
        return "search for mobile" in text or "global search" in text or "groups chats" in text

    def search_input_element(self, search_key: str) -> Any | None:
        window = self.focus_window()
        for element in window.descendants():
            try:
                control_type = str(element.element_info.control_type or "")
                class_name = str(element.element_info.class_name or "")
                text = element.window_text().strip()
            except Exception:
                continue

            is_edit = control_type == "Edit" or "Edit" in class_name
            if is_edit and text == search_key:
                return element
        return None

    def search_input_contains(self, search_key: str) -> bool:
        return self.search_input_element(search_key) is not None

    def search_is_active(self, job: PasteJob) -> bool:
        return self.search_overlay_visible() or self.search_input_contains(job.search_key)

    def clear_search_state(self, job: PasteJob) -> None:
        try:
            search_input = self.search_input_element(job.search_key)
            if search_input is not None:
                try:
                    search_input.set_focus()
                    time.sleep(0.1)
                except Exception:
                    pass
            else:
                self.focus_window()
                self.send_keys(self.search_shortcut)
                time.sleep(0.2)
            self.send_keys("^a")
            time.sleep(0.05)
            self.send_keys("{BACKSPACE}")
            time.sleep(0.1)
            self.send_keys("{ESC}")
            time.sleep(self.settle_seconds)
        except Exception:
            pass

    def open_best_result_with_uia(self, job: PasteJob) -> bool:
        for _, element, text in self.search_result_candidates(job):
            print(f"Trying UI Automation result: {text[:80]}")
            try:
                element.invoke()
                time.sleep(self.settle_seconds)
                if not self.search_is_active(job):
                    return True
            except Exception:
                pass

        return False

    def open_first_result_with_keyboard(self) -> None:
        self.send_keys("{DOWN}")
        time.sleep(0.15)
        self.send_keys("{ENTER}")
        time.sleep(self.settle_seconds)

    def print_search_result_candidates(self, job: PasteJob) -> None:
        self.search_chat(job.search_key)
        candidates = self.search_result_candidates(job)
        if not candidates:
            print("No safe group-chat candidates found above the online/global-search rows.")
            return
        print("Safe WeCom search candidates:")
        for score, element, text in candidates[:10]:
            print(f"- score={score}; text={text[:120]}")

    def open_chat_from_search(self, job: PasteJob, *, open_strategy: str = "enter-first") -> None:
        self.search_chat(job.search_key)

        if open_strategy == "manual-click":
            input("Click the correct WeCom search result/chat, then press Enter here to close search and verify. ")
            self.send_keys("{ESC}")
            time.sleep(self.settle_seconds)
            return

        if open_strategy == "enter":
            self.send_keys("{ENTER}")
            time.sleep(self.settle_seconds)
            if self.search_is_active(job):
                self.clear_search_state(job)
                raise LookupError(f"WeCom group chat was not found for uid {job.uid}.")
            return

        if open_strategy == "keyboard":
            self.open_first_result_with_keyboard()
        elif open_strategy == "ui-control":
            if not self.open_best_result_with_uia(job):
                self.clear_search_state(job)
                raise LookupError(f"WeCom group chat was not found for uid {job.uid}.")
        else:
            print("Opening first WeCom search result with Enter.")
            self.send_keys("{ENTER}")
            time.sleep(self.settle_seconds)
            if self.search_is_active(job):
                print("Enter did not open the result; using UI Automation fallback.")
                if not self.open_best_result_with_uia(job):
                    self.clear_search_state(job)
                    raise LookupError(f"WeCom group chat was not found for uid {job.uid}.")

        if self.search_is_active(job):
            self.clear_search_state(job)
            raise LookupError(f"WeCom group chat was not found for uid {job.uid}.")
        time.sleep(self.settle_seconds)

    def verify_chat(self, job: PasteJob) -> tuple[bool, str]:
        text = self.visible_text()
        uid_ok = job.uid in text
        name_parts = [part for part in job.student_name.split() if part]
        name_ok = any(part in text for part in name_parts)

        if uid_ok and name_ok:
            return True, "verified_uid_and_name"
        if uid_ok:
            return True, "verified_uid_only"
        return False, "uid_not_visible_after_search"

    def print_visible_text_debug(self, *, limit: int = 2000) -> None:
        text = self.visible_text()
        print("Visible WeCom text sample:")
        print(text[:limit] if text else "(no visible text captured)")

    def focus_message_input(self) -> bool:
        window = self.focus_window()
        candidates: list[tuple[int, Any]] = []

        for element in window.descendants():
            try:
                rectangle = element.rectangle()
                control_type = str(element.element_info.control_type or "")
                class_name = str(element.element_info.class_name or "")
                text = element.window_text().strip()
            except Exception:
                continue

            is_edit = control_type == "Edit" or "Edit" in class_name
            if not is_edit:
                continue
            if text and text == self.search_shortcut:
                continue
            if rectangle.width() < 100 or rectangle.height() < 20:
                continue

            # The message composer is usually the lowest sizeable edit control.
            candidates.append((rectangle.top, element))

        if candidates:
            _, element = max(candidates, key=lambda candidate: candidate[0])
            try:
                element.set_focus()
                time.sleep(0.2)
                return True
            except Exception:
                pass

        return False

    def paste_feedback(self, feedback: str) -> None:
        try:
            import pyperclip
        except ImportError as exc:
            raise RuntimeError("Install clipboard dependency first: python -m pip install pyperclip") from exc

        if not self.focus_message_input():
            print("Could not focus the WeCom message input box by UI Automation; pasting directly into the active WeCom chat.")
            self.focus_window()

        pyperclip.copy(feedback)
        self.send_keys("^v")
        time.sleep(0.8)


class WhatsAppPasteRobot:
    def __init__(
        self,
        *,
        title_re: str = DEFAULT_WHATSAPP_TITLE_RE,
        search_shortcut: str = "^f",
        settle_seconds: float = 1.0,
    ) -> None:
        try:
            from pywinauto import Desktop
            from pywinauto.keyboard import send_keys
        except ImportError as exc:
            raise RuntimeError(
                "Install desktop automation dependencies first: "
                "python -m pip install pywinauto pyperclip"
            ) from exc

        self.desktop = Desktop(backend="uia")
        self.send_keys = send_keys
        self.title_re = title_re
        self.search_shortcut = search_shortcut
        self.settle_seconds = settle_seconds
        self.window: Any | None = None

    def focus_window(self) -> Any:
        if self.window is not None:
            try:
                if self.window.exists(timeout=0.5):
                    self.window.set_focus()
                    return self.window
            except Exception:
                self.window = None

        window = find_window_by_title_re(self.title_re)
        if window is None:
            raise RuntimeError(f"Could not find WhatsApp window with title pattern: {self.title_re}")
        window.set_focus()
        self.window = window
        return window

    def visible_text(self) -> str:
        window = self.focus_window()
        texts: list[str] = []
        try:
            title = window.window_text().strip()
        except Exception:
            title = ""
        if title:
            texts.append(title)
        for element in window.descendants():
            try:
                text = element.window_text().strip()
            except Exception:
                continue
            if text:
                texts.append(text)
        return "\n".join(texts)

    def open_phone_url(self, job: PasteJob) -> None:
        if not job.whatsapp_phone:
            raise RuntimeError("WhatsApp phone mode needs a WhatsApp Phone value.")

        url = f"https://wa.me/{job.whatsapp_phone}?text={quote(job.feedback)}"
        os.startfile(url)  # type: ignore[attr-defined]
        time.sleep(3)

    def search_chat(self, search_key: str) -> None:
        try:
            import pyperclip
        except ImportError as exc:
            raise RuntimeError("Install clipboard dependency first: python -m pip install pyperclip") from exc

        self.focus_window()
        self.send_keys("{ESC}")
        time.sleep(0.1)
        pyperclip.copy(search_key)
        self.send_keys(self.search_shortcut)
        time.sleep(0.3)
        self.send_keys("^a")
        self.send_keys("^v")
        time.sleep(self.settle_seconds)

    def search_input_element(self, search_key: str) -> Any | None:
        window = self.focus_window()
        for element in window.descendants():
            try:
                control_type = str(element.element_info.control_type or "")
                class_name = str(element.element_info.class_name or "")
                text = element.window_text().strip()
            except Exception:
                continue

            is_edit = control_type == "Edit" or "Edit" in class_name
            if is_edit and text == search_key:
                return element
        return None

    def search_input_contains(self, search_key: str) -> bool:
        return self.search_input_element(search_key) is not None

    def element_has_keyboard_focus(self, element: Any) -> bool:
        try:
            return bool(element.has_keyboard_focus())
        except Exception:
            pass
        try:
            return bool(element.element_info.has_keyboard_focus)
        except Exception:
            return False

    def search_is_active(self, search_key: str) -> bool:
        search_input = self.search_input_element(search_key)
        return search_input is not None and self.element_has_keyboard_focus(search_input)

    def clear_search_state(self, search_key: str) -> None:
        try:
            self.focus_window()
            self.send_keys(self.search_shortcut)
            time.sleep(0.2)
            self.send_keys("^a")
            time.sleep(0.05)
            self.send_keys("{BACKSPACE}")
            time.sleep(0.1)
            self.send_keys("{ESC}")
            time.sleep(self.settle_seconds)
        except Exception:
            pass

    def open_chat_from_search(self, search_key: str) -> None:
        self.search_chat(search_key)
        print("Opening WhatsApp search result with Enter.")
        self.send_keys("{ENTER}")
        time.sleep(self.settle_seconds)
        if self.search_is_active(search_key):
            self.clear_search_state(search_key)
            raise LookupError(f"WhatsApp chat was not found for search key {search_key!r}.")

    def close_search_overlay(self) -> None:
        try:
            self.send_keys("{ESC}")
            time.sleep(0.2)
        except Exception:
            pass

    def verify_chat(self, job: PasteJob) -> tuple[bool, str]:
        text = self.visible_text()
        uid_ok = bool(job.uid and job.uid in text)
        search_ok = bool(job.search_key and job.search_key in text)
        name_parts = [part for part in job.student_name.split() if part]
        name_ok = any(part in text for part in name_parts)

        if uid_ok or search_ok:
            return True, "verified_search_key"
        if name_ok:
            return True, "verified_name_only"
        return False, "whatsapp_target_not_visible"

    def paste_feedback(self, feedback: str) -> None:
        try:
            import pyperclip
        except ImportError as exc:
            raise RuntimeError("Install clipboard dependency first: python -m pip install pyperclip") from exc

        for attempt in range(3):
            pyperclip.copy(feedback)
            time.sleep(0.2)
            if pyperclip.paste() == feedback:
                print(f"Clipboard loaded with feedback ({len(feedback)} characters).")
                break
            if attempt == 2:
                raise RuntimeError("Could not copy feedback text into the clipboard.")

        self.send_keys("^v")
        time.sleep(1.5)


def print_job(job: PasteJob) -> None:
    print("=" * 72)
    print(f"Row: {job.excel_row}")
    print(f"UID: {job.uid}")
    print(f"Student: {job.student_name}")
    print(f"Parent language: {job.parent_language}")
    print(f"Channel: {job.channel}")
    print(f"Search key: {job.search_key}")
    print(f"Expected chat name: {job.expected_chat_name}")
    if job.channel == "whatsapp":
        print(f"WhatsApp target type: {job.whatsapp_target_type}")
        print(f"WhatsApp phone: {job.whatsapp_phone or '(blank)'}")
    print()
    print(job.feedback)
    print()


def configured_exe_for_app(args: argparse.Namespace, app_key: str) -> Path | None:
    if app_key == "wecom":
        return args.wecom_exe
    if app_key == "whatsapp":
        return args.whatsapp_exe
    return None


def print_readiness(job: PasteJob, status: DesktopAppStatus) -> None:
    print_app_status(status)
    ready = status.dependency_ok and status.window_found
    if job.channel == "whatsapp" and job.whatsapp_target_type == "phone":
        print("Safe to run paste-only for this row: yes, phone URL mode does not need app window verification")
    else:
        print(f"Safe to run paste-only for this row: {'yes' if ready else 'no'}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Supervised paste helper for WeCom / 企业微信 and WhatsApp feedback messages. It never sends."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--row", type=int, default=2)
    parser.add_argument(
        "--rows",
        help="Comma-separated rows or ranges to process, such as 3,5 or 3-5.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        help="First row for a batch range. Use with --end-row.",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        help="Last row for a batch range.",
    )
    parser.add_argument("--class-review", default="")
    parser.add_argument("--class-review-file", type=Path)
    parser.add_argument("--use-api", action="store_true")
    parser.add_argument("--model", default=DEFAULT_OPENAI_MODEL)
    parser.add_argument(
        "--status",
        action="store_true",
        help="Check workbook row, needed app, and desktop automation readiness without pasting.",
    )
    parser.add_argument(
        "--open-app",
        action="store_true",
        help="Open the app needed for this row if it is not already available, then stop unless a paste mode is selected.",
    )
    parser.add_argument(
        "--debug-search-results",
        action="store_true",
        help="Search WeCom / 企业微信 and print safe candidate UI elements without opening, pasting, or sending.",
    )
    parser.add_argument(
        "--debug-window-titles",
        action="store_true",
        help="Print app window titles that match the selected channel without pasting.",
    )
    parser.add_argument(
        "--mode",
        choices=["dry-run", "paste-only"],
        default="dry-run",
        help="dry-run prints the job only; paste-only searches the selected app, verifies when possible, and pastes without sending.",
    )
    parser.add_argument(
        "--no-auto-open",
        action="store_true",
        help="In paste-only mode, do not try to launch the needed app automatically.",
    )
    parser.add_argument(
        "--auto-open-search-result",
        action="store_true",
        help="After searching WeCom / 企业微信, press Enter only.",
    )
    parser.add_argument(
        "--ui-control-result-open",
        action="store_true",
        help="After searching WeCom / 企业微信, open a matching result with UI Automation.",
    )
    parser.add_argument(
        "--keyboard-result-open",
        action="store_true",
        help="After searching WeCom / 企业微信, use Down then Enter to open the first result.",
    )
    parser.add_argument(
        "--manual-result-click",
        action="store_true",
        help="After searching WeCom / 企业微信, wait for you to click the search result.",
    )
    parser.add_argument(
        "--require-verification",
        action="store_true",
        help="Stop before pasting if the script cannot verify the WeCom chat by visible UI text.",
    )
    parser.add_argument("--wecom-title-re", default=DEFAULT_WECOM_TITLE_RE)
    parser.add_argument("--whatsapp-title-re", default=DEFAULT_WHATSAPP_TITLE_RE)
    parser.add_argument("--wecom-exe", type=Path, help="Optional explicit path to WXWork.exe / WeCom.exe.")
    parser.add_argument("--whatsapp-exe", type=Path, help="Optional explicit path to WhatsApp.exe.")
    parser.add_argument("--search-shortcut", default="^f")
    parser.add_argument(
        "--whatsapp-search-shortcut",
        default="^f",
        help="Keyboard shortcut used to open WhatsApp search. Default is Ctrl+F.",
    )
    parser.add_argument(
        "--no-status-write",
        action="store_true",
        help="Do not write Send Status, Send Error, or Last Attempt back to the workbook.",
    )
    return parser


def open_strategy_from_args(args: argparse.Namespace) -> str:
    if args.manual_result_click:
        return "manual-click"
    if args.auto_open_search_result:
        return "enter"
    if args.ui_control_result_open:
        return "ui-control"
    if args.keyboard_result_open:
        return "keyboard"
    return "enter-first"


def ensure_desktop_ready(
    app_spec: DesktopAppSpec,
    *,
    app_exe: Path | None,
    no_auto_open: bool,
) -> DesktopAppStatus:
    status = ensure_app_available(
        app_spec,
        configured_path=app_exe,
        open_if_missing=not no_auto_open,
    )
    print_app_status(status)
    if not status.dependency_ok:
        raise RuntimeError(
            "Desktop automation dependencies are not available. "
            "Run: python -m pip install -r requirements.txt"
        )
    if not status.window_found:
        raise RuntimeError("Needed app window is not available.")
    return status


def run_wecom_job(
    job: PasteJob,
    *,
    args: argparse.Namespace,
    app_spec: DesktopAppSpec,
    app_exe: Path | None,
) -> JobResult:
    ensure_desktop_ready(app_spec, app_exe=app_exe, no_auto_open=args.no_auto_open)
    robot = WeComPasteRobot(
        title_re=args.wecom_title_re,
        search_shortcut=args.search_shortcut,
    )

    if args.debug_search_results:
        robot.print_search_result_candidates(job)
        return JobResult(status="checked")

    try:
        robot.open_chat_from_search(
            job,
            open_strategy=open_strategy_from_args(args),
        )
    except Exception:
        robot.clear_search_state(job)
        raise

    if robot.search_is_active(job):
        robot.clear_search_state(job)
        raise LookupError(f"WeCom group chat was not found for uid {job.uid}.")

    verified, reason = robot.verify_chat(job)
    print(f"Verification: {reason}")
    if not verified:
        robot.print_visible_text_debug()
        if args.require_verification:
            raise LookupError("WeCom chat could not be verified.")
        print("WARNING: Could not verify the chat automatically. Continuing because paste-only does not send.")

    robot.paste_feedback(job.feedback)
    robot.clear_search_state(job)
    print("Cleared WeCom search box for the next row.")
    return JobResult(status="pasted", pasted=True)


def run_whatsapp_job(
    job: PasteJob,
    *,
    args: argparse.Namespace,
    app_spec: DesktopAppSpec,
    app_exe: Path | None,
) -> JobResult:
    if job.whatsapp_target_type == "phone":
        if not job.whatsapp_phone:
            raise LookupError("WhatsApp phone mode needs a WhatsApp Phone value.")
        url = f"https://wa.me/{job.whatsapp_phone}?text={quote(job.feedback)}"
        os.startfile(url)  # type: ignore[attr-defined]
        time.sleep(3)
        return JobResult(status="pasted", pasted=True)

    if not job.search_key:
        raise LookupError("WhatsApp group_search needs WhatsApp Search Key or uid.")

    ensure_desktop_ready(app_spec, app_exe=app_exe, no_auto_open=args.no_auto_open)
    robot = WhatsAppPasteRobot(
        title_re=args.whatsapp_title_re,
        search_shortcut=args.whatsapp_search_shortcut,
    )
    try:
        robot.open_chat_from_search(job.search_key)
    except Exception:
        robot.clear_search_state(job.search_key)
        raise

    robot.paste_feedback(job.feedback)

    verified, reason = robot.verify_chat(job)
    print(f"Verification: {reason}")
    if not verified:
        if args.require_verification:
            raise LookupError("WhatsApp chat could not be verified.")
        print("WARNING: Could not verify the WhatsApp chat automatically. Continuing because paste-only does not send.")

    robot.clear_search_state(job.search_key)
    print("Cleared WhatsApp search box for the next row.")

    return JobResult(status="pasted", pasted=True)


def run_job(
    job: PasteJob,
    *,
    args: argparse.Namespace,
    app_specs: dict[str, DesktopAppSpec],
    batch_mode: bool,
) -> JobResult:
    print_job(job)

    if job.channel not in app_specs:
        return JobResult(status="needs_review", error=f"No desktop app setup is defined for channel {job.channel!r}.")

    app_spec = app_specs[job.channel]
    app_exe = configured_exe_for_app(args, job.channel)

    if args.status:
        status = app_status(app_spec)
        print_readiness(job, status)
        return JobResult(status="checked")

    if args.debug_window_titles:
        print_matching_window_titles(app_spec.title_re)
        return JobResult(status="checked")

    if args.open_app:
        status = ensure_app_available(
            app_spec,
            configured_path=app_exe,
            open_if_missing=True,
        )
        print_readiness(job, status)
        if args.mode == "dry-run":
            return JobResult(status="checked")

    if args.mode == "dry-run" and not args.debug_search_results:
        print("Dry run only. Nothing was pasted or sent.")
        return JobResult(status="dry_run")

    clear_clipboard()
    try:
        if job.channel == "wecom":
            result = run_wecom_job(job, args=args, app_spec=app_spec, app_exe=app_exe)
        elif job.channel == "whatsapp":
            result = run_whatsapp_job(job, args=args, app_spec=app_spec, app_exe=app_exe)
        else:
            result = JobResult(status="needs_review", error=f"Unsupported channel {job.channel!r}.")
    except LookupError as exc:
        result = JobResult(status="needs_review", error=str(exc))
    except Exception as exc:
        result = JobResult(status="failed", error=f"{type(exc).__name__}: {exc}")
    finally:
        time.sleep(0.2)
        clear_clipboard()

    if result.pasted:
        print("Pasted only. The script did not send the message.")
    elif result.error:
        print(f"{result.status}: {result.error}")
    return result


def main() -> None:
    args = build_parser().parse_args()
    class_review = args.class_review
    if args.class_review_file:
        class_review = args.class_review_file.read_text(encoding="utf-8").strip()

    jobs = load_jobs(
        args.workbook,
        sheet_name=args.sheet,
        row_numbers=selected_row_numbers(args),
        class_review=class_review,
        use_api=args.use_api,
        model=args.model,
    )

    app_specs = build_app_specs(
        wecom_title_re=args.wecom_title_re,
        whatsapp_title_re=args.whatsapp_title_re,
    )

    batch_mode = len(jobs) > 1
    should_write_status = args.mode == "paste-only" and not args.no_status_write
    if should_write_status:
        ensure_status_columns(args.workbook, args.sheet)

    processed = 0
    pasted = 0
    needs_review = 0
    skipped = 0
    failed = 0
    for index, item in enumerate(jobs, start=1):
        if batch_mode:
            print(f"Batch item {index}/{len(jobs)}")

        if isinstance(item, JobLoadError):
            print("=" * 72)
            print(f"Row: {item.row_number}")
            print(f"{item.status}: {item.error}")
            if should_write_status:
                write_job_status(
                    args.workbook,
                    args.sheet,
                    item.row_number,
                    status=item.status,
                    error=item.error,
                )
            if item.status == "skipped_absent":
                skipped += 1
            else:
                needs_review += 1
            clear_clipboard()
            continue

        result = run_job(item, args=args, app_specs=app_specs, batch_mode=batch_mode)
        processed += 1
        if result.pasted:
            pasted += 1
        elif result.status == "needs_review":
            needs_review += 1
        elif result.status == "skipped_absent":
            skipped += 1
        elif result.status == "failed":
            failed += 1

        if should_write_status and result.status in {"pasted", "needs_review", "skipped_absent", "failed"}:
            write_job_status(
                args.workbook,
                args.sheet,
                item.excel_row,
                status=result.status,
                error=result.error,
            )

    if batch_mode:
        print("=" * 72)
        print(
            "Batch complete. "
            f"Processed: {processed}. Pasted: {pasted}. "
            f"Needs review: {needs_review}. Skipped: {skipped}. Failed: {failed}. Sent: 0."
        )


if __name__ == "__main__":
    main()
