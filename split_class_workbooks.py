from __future__ import annotations

import argparse
from pathlib import Path
import re

from openpyxl import load_workbook

from feedback_generator import DEFAULT_WORKBOOK


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = PROJECT_DIR / "class_workbooks"


def safe_filename(value: str) -> str:
    text = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", value.strip(), flags=re.UNICODE)
    return text.strip("_") or "class"


def split_workbook(source_path: Path, output_dir: Path) -> list[Path]:
    source = load_workbook(source_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    created: list[Path] = []

    for sheet_name in source.sheetnames:
        class_workbook = load_workbook(source_path)
        for other_sheet in list(class_workbook.sheetnames):
            if other_sheet != sheet_name:
                del class_workbook[other_sheet]

        class_workbook.active.title = sheet_name
        output_path = output_dir / f"{safe_filename(sheet_name)}.xlsx"
        class_workbook.save(output_path)
        created.append(output_path)

    return created


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Split the combined feedback workbook into one workbook per class.")
    parser.add_argument("--source", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    created = split_workbook(args.source, args.output_dir)
    for path in created:
        print(path)
    print(f"Created {len(created)} class workbook(s).")


if __name__ == "__main__":
    main()
