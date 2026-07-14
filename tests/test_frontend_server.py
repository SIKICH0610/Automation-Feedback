from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from database_store import StoreError
from feedback_generator import parse_row_numbers
from frontend_server import ActionRunner, WorkbookStore, announcement_filename


HEADERS = [
    "First Name",
    "Last Name",
    "uid",
    "Parent Language",
    "Remark for Student",
    "Feedback",
    "Quiz1 Feedback",
    "Quiz1 Mistake",
    "Quiz1 Score",
    "Quiz1 Average",
    "Quiz2 Feedback",
    "Quiz2 Mistake",
    "Quiz2 Score",
    "Quiz2 Average",
]


class FrontendStoreTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory(dir=Path(__file__).parent)
        root = Path(self.temp_dir.name)
        self.source_workbook_path = root / "students.xlsx"
        self.app_data_dir = root / "app_data"
        self.database_path = self.app_data_dir / "feedback.db"
        self.export_dir = root / "exports"
        self.announcement_dir = root / "announcements"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Class A"
        sheet.append(HEADERS)
        sheet.append(
            [
                "Ada",
                "Lovelace",
                "1001",
                "English",
                "Focused",
                "",
                "",
                "",
                "8/10",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
        workbook.create_sheet("Class B").append(HEADERS)
        workbook.save(self.source_workbook_path)
        workbook.close()

        self.store = self.make_store()

    def make_store(self) -> WorkbookStore:
        return WorkbookStore(
            database_path=self.database_path,
            source_workbook_path=self.source_workbook_path,
            announcement_dir=self.announcement_dir,
            app_data_dir=self.app_data_dir,
            export_dir=self.export_dir,
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_first_run_imports_sheets_and_linked_announcement(self) -> None:
        self.assertTrue(self.database_path.exists())
        self.assertEqual(self.store.sheet_names(), ["Class A", "Class B"])
        data = self.store.load_sheet("Class A")
        self.assertEqual(data["rows"][0]["values"]["First Name"], "Ada")
        self.assertTrue(data["rows"][0]["student_id"])
        self.assertEqual(Path(data["announcement_path"]).name, "Class_A.txt")
        self.assertTrue(Path(data["announcement_path"]).exists())

    def test_database_save_and_export_leave_source_workbook_unchanged(self) -> None:
        data = self.store.load_sheet("Class A")
        data["rows"][0]["values"]["Remark for Student"] = "Strong proof work"
        data["rows"].append(
            {
                "student_id": None,
                "values": {
                    header: (
                        "Grace"
                        if header == "First Name"
                        else "Hopper"
                        if header == "Last Name"
                        else "1002"
                        if header == "uid"
                        else ""
                    )
                    for header in HEADERS
                },
            }
        )

        saved = self.store.save_sheet("Class A", data["rows"])
        self.assertEqual(len(saved["rows"]), 2)
        self.assertEqual(saved["rows"][0]["values"]["Remark for Student"], "Strong proof work")

        source = load_workbook(self.source_workbook_path)
        self.assertEqual(source["Class A"]["E2"].value, "Focused")
        self.assertIsNone(source["Class A"]["A3"].value)
        source.close()

        export_path = self.store.export_public_workbook()
        exported = load_workbook(export_path)
        sheet = exported["Class A"]
        self.assertEqual(sheet["E2"].value, "Strong proof work")
        self.assertEqual(sheet["A3"].value, "Grace")
        self.assertEqual(sheet["A2"].style_id, sheet["A3"].style_id)
        exported.close()

    def test_original_excel_edits_are_ignored_after_migration(self) -> None:
        original_id = self.store.load_sheet("Class A")["rows"][0]["student_id"]

        workbook = load_workbook(self.source_workbook_path)
        workbook["Class A"]["A2"] = "Changed outside the app"
        workbook.save(self.source_workbook_path)
        workbook.close()
        self.store.template_path.unlink()

        reopened = self.make_store()
        row = reopened.load_sheet("Class A")["rows"][0]
        self.assertEqual(row["student_id"], original_id)
        self.assertEqual(row["values"]["First Name"], "Ada")
        self.assertTrue(reopened.template_path.exists())

        export_path = reopened.export_public_workbook()
        exported = load_workbook(export_path)
        self.assertEqual(exported["Class A"]["A2"].value, "Ada")
        exported.close()

    def test_runtime_workbook_changes_sync_back_to_database(self) -> None:
        self.store.prepare_runtime_workbook()
        runtime = load_workbook(self.store.workbook_path)
        sheet = runtime["Class A"]
        feedback_column = HEADERS.index("Feedback") + 1
        sheet.cell(2, feedback_column).value = "Generated feedback"
        runtime.save(self.store.workbook_path)
        runtime.close()

        self.store.sync_runtime_workbook()
        self.store.remove_runtime_workbook()

        data = self.store.load_sheet("Class A")
        self.assertEqual(data["rows"][0]["values"]["Feedback"], "Generated feedback")
        self.assertFalse(self.store.workbook_path.exists())

    def test_incomplete_or_duplicate_roster_save_is_rejected(self) -> None:
        data = self.store.load_sheet("Class A")
        with self.assertRaises(StoreError):
            self.store.save_sheet("Class A", [])

        duplicate = [data["rows"][0], data["rows"][0]]
        with self.assertRaises(StoreError):
            self.store.save_sheet("Class A", duplicate)

    def test_search_students_matches_name_uid_and_class(self) -> None:
        class_b = self.store.load_sheet("Class B")
        class_b["rows"].append(
            {
                "student_id": None,
                "values": {
                    header: (
                        "Grace"
                        if header == "First Name"
                        else "Hopper"
                        if header == "Last Name"
                        else "2002"
                        if header == "uid"
                        else ""
                    )
                    for header in HEADERS
                },
            }
        )
        self.store.save_sheet("Class B", class_b["rows"])

        by_name = self.store.search_students("grace")
        self.assertEqual(
            [(row["full_name"], row["sheet"]) for row in by_name],
            [("Grace Hopper", "Class B")],
        )
        by_uid = self.store.search_students("1001")
        self.assertEqual(by_uid[0]["full_name"], "Ada Lovelace")
        self.assertEqual(by_uid[0]["sheet"], "Class A")
    def test_action_commands_use_runtime_workbook_and_requested_columns(self) -> None:
        runner = ActionRunner(self.store)
        command, _ = runner.command_for(
            {
                "action": "generate-quiz-feedback",
                "sheet": "Class A",
                "rows": [2],
                "quiz_number": 1,
            }
        )
        joined = " ".join(command)
        self.assertIn(str(self.store.workbook_path), joined)
        self.assertIn("Quiz1 Feedback", joined)
        self.assertIn("Quiz1 Score", joined)
        self.assertIn("Quiz1 Average", joined)

        paste_command, _ = runner.command_for(
            {
                "action": "paste-comments",
                "sheet": "Class A",
                "rows": [2],
            }
        )
        self.assertIn("paste_sender.py", " ".join(paste_command))
        self.assertIn("paste-only", paste_command)

    def test_row_parser_and_safe_announcement_name(self) -> None:
        self.assertEqual(parse_row_numbers("2,5,7-9,5"), [2, 5, 7, 8, 9])
        self.assertEqual(announcement_filename("Geo TTh"), "Geo_TTh.txt")


if __name__ == "__main__":
    unittest.main()


