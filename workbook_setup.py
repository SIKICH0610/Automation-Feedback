from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from feedback_generator import ADDITIONAL_COMMENT_COLUMN, DEFAULT_WORKBOOK


DEFAULT_COLUMNS = [
    (ADDITIONAL_COMMENT_COLUMN, "Homework Reflection"),
    ("Preferred Channel", "Send Status"),
    ("WhatsApp Phone", "Send Status"),
    ("WhatsApp Search Key", "Send Status"),
    ("WhatsApp Target Type", "Send Status"),
    ("Send Error", "Send Status"),
    ("Last Attempt", "Send Error"),
]


def header_values(worksheet: Any) -> list[str]:
    return [
        str(worksheet.cell(1, col).value or "").strip()
        for col in range(1, worksheet.max_column + 1)
    ]


def ensure_column(workbook_path: Path, column_name: str, before_column: str) -> list[str]:
    workbook = load_workbook(workbook_path)
    changed_sheets: list[str] = []

    for worksheet in workbook.worksheets:
        headers = header_values(worksheet)
        if column_name in headers:
            continue

        if before_column in headers:
            insert_at = headers.index(before_column) + 1
            worksheet.insert_cols(insert_at)
        else:
            insert_at = worksheet.max_column + 1

        worksheet.cell(1, insert_at).value = column_name
        changed_sheets.append(worksheet.title)

    if changed_sheets:
        workbook.save(workbook_path)
    return changed_sheets


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Prepare workbook columns for feedback automation.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument(
        "--column",
        help="Optional single column to add. Omit to add all automation helper columns.",
    )
    parser.add_argument(
        "--before-column",
        default="Homework Reflection",
        help="Insert the new column before this header when possible.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.column:
        changed_sheets = ensure_column(args.workbook, args.column, args.before_column)
        if changed_sheets:
            print(f"Added {args.column!r} to: {', '.join(changed_sheets)}")
        else:
            print(f"{args.column!r} already exists on every sheet.")
        return

    any_changes = False
    for column_name, before_column in DEFAULT_COLUMNS:
        changed_sheets = ensure_column(args.workbook, column_name, before_column)
        if changed_sheets:
            any_changes = True
            print(f"Added {column_name!r} to: {', '.join(changed_sheets)}")

    if not any_changes:
        print("Automation helper columns already exist on every sheet.")


if __name__ == "__main__":
    main()
